#!/usr/bin/env python3
"""
Document Classification & Extraction System v3 for Romanian Construction Documents.

Classification: 3-level approach:
  Level 1: Filename regex (fast, deterministic)
  Level 2: Text content analysis (rule-based, no AI)
  Level 3: AI via OpenRouter (fallback for scanned/ambiguous docs)

Extraction: AI-powered structured data extraction per category:
  - Uses category-specific extraction schemas
  - Multi-model support for testing/comparison
  - OCR fallback for scanned documents
  - Outputs: Nume Document, Numar Pagini, Material, Data Expirare,
             Companie, Producator, Distribuitor, Adresa Producator
"""

import os
import re
import sys
import json
import time
import io
import warnings
import logging
import pdfplumber
import fitz  # PyMuPDF
import requests
import pytesseract
from PIL import Image
from pathlib import Path
from collections import defaultdict

# Auto-detect Tesseract on Windows (common install paths)
import shutil
if not shutil.which("tesseract"):
    for _tess_path in [
        r"D:\Tesseract-OCR\tesseract.exe",
        r"E:\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]:
        if os.path.isfile(_tess_path):
            pytesseract.pytesseract.tesseract_cmd = _tess_path
            break

# Suppress noisy PDF parsing warnings
warnings.filterwarnings("ignore", message="Cannot set gray non-stroke color")
logging.getLogger("pdfminer").setLevel(logging.ERROR)

# ============================================================
# CONFIG
# ============================================================
OPENROUTER_API_KEY = "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"
AI_MODEL = "google/gemini-2.0-flash-001"
EXTRACTION_MODEL = "google/gemini-2.0-flash-001"  # Can be overridden via CLI
BASE_DIR = "/sessions/sharp-cool-planck/mnt/makyol-automation--makyol-automation/Fisiere Makyol Clasificate"

FOLDER_TO_CATEGORY = {
    "Agrement": "AGREMENT",
    "Autorizatii de comercializare sau de distributie": "AUTORIZATIE_DISTRIBUTIE",
    "Aviz Tehnic + Agrement": "AVIZ_TEHNIC_SI_AGREMENT",
    "Aviz tehnic": "AVIZ_TEHNIC",
    "Avize Sanitare": "AVIZ_SANITAR",
    "CE": "CE",
    "Certificat Calitate": "CERTIFICAT_CALITATE",
    "Certificat Garantie": "CERTIFICAT_GARANTIE",
    "Cui": "CUI",
    "Declaratie de Performanta": "DECLARATIE_PERFORMANTA",
    "Declaratii conformitate": "DECLARATIE_CONFORMITATE",
    "Fisa Tehnica": "FISA_TEHNICA",
    "Iso": "ISO",
    "Z.Altele": "ALTELE",
}


# ============================================================
# DATA EXTRACTION SCHEMAS (per category)
# ============================================================

EXTRACTION_SCHEMA = {
    "AUTORIZATIE_DISTRIBUTIE": {
        "fields": ["distribuitor", "producator", "data_emitere", "valabilitate", "data_expirare", "material", "adresa_producator"],
        "instructions": "Extract: distributor company, producer company, issue date, validity period, expiration date, material/product, producer address. The material is the product authorized for distribution (e.g., 'tevi si fitinguri PEID', 'vane industriale'). If the authorization is generic ('produsele achizitionate'), write 'Produse diverse (autorizatie generala)'. If no explicit expiration date but has validity period, CALCULATE it. If validity = 'pe durata contractului', write 'Pe durata contractului'."
    },
    "AGREMENT": {
        "fields": ["producator", "data_expirare", "material", "companie", "adresa_producator"],
        "instructions": "Extract: producer name, expiration date, material/product name, company that holds the agrement, producer address. Look for 'valabil pana la', 'valabilitate' for the expiration date."
    },
    "AVIZ_TEHNIC_SI_AGREMENT": {
        "fields": ["producator", "data_expirare", "material", "companie", "adresa_producator"],
        "instructions": "Extract: producer name, expiration date, material/product, company name, producer address. For 'nume_document', this should be 'Aviz Tehnic si Agrement Tehnic' (NOT just 'Aviz Tehnic')."
    },
    "AVIZ_TEHNIC": {
        "fields": ["producator", "data_expirare", "material", "companie", "adresa_producator"],
        "instructions": "Extract: producer name, expiration date, material/product, company (the certification body if applicable), producer address."
    },
    "AVIZ_SANITAR": {
        "fields": ["companie", "material", "data_expirare", "producator", "adresa_producator"],
        "instructions": "Extract: company holding the sanitary approval, material/product, expiration date, producer, producer address. Look very carefully for expiration - check for 'valabil', 'expira', 'pana la', and also check if the approval number contains a year-based validity. If no expiration is stated anywhere, use null."
    },
    "ISO": {
        "fields": ["companie", "data_expirare", "standard_iso", "adresa_producator"],
        "instructions": "Extract: certified company name, certificate expiration date, ISO standard number (e.g., 'ISO 9001:2015'), company address. NOTE: Do NOT put the ISO standard in the 'material' field - ISO certifies management systems, not physical materials. For 'nume_document', write 'Certificat ISO [standard]' (e.g., 'Certificat ISO 9001')."
    },
    "DECLARATIE_PERFORMANTA": {
        "fields": ["material", "producator", "adresa_producator"],
        "instructions": "Extract: product/material name (full description with specs), producer name, producer address. Keep material name concise but specific (max 100 chars)."
    },
    "FISA_TEHNICA": {
        "fields": ["producator", "material", "adresa_producator"],
        "instructions": "Extract: producer name, product/material name (full name with type, e.g., 'Tevi PE100 PEHD pentru apa'), producer address."
    },
    "CERTIFICAT_GARANTIE": {
        "fields": ["material", "producator", "adresa_producator", "data_expirare"],
        "instructions": "Extract: product/material under warranty (keep CONCISE, max 80 chars - summarize if the list is long, e.g., 'Tevi si fitinguri PVC, PP, PE, PEX'), producer name, producer address, warranty period/expiration. For data_expirare: extract the warranty DURATION (e.g., '2 ani de la receptie', '24 luni'). If multiple warranty periods exist, extract the main/longest one."
    },
    "DECLARATIE_CONFORMITATE": {
        "fields": ["material", "producator", "companie", "adresa_producator"],
        "instructions": "Extract: product/material name (the actual product - if generic like 'produse pentru constructii', keep it as is but add any specifics from the document context), producer name, declaring company, producer address."
    },
    "CE": {
        "fields": ["producator", "companie", "material", "data_expirare", "adresa_producator"],
        "instructions": "Extract data from this CE/PED certificate. IMPORTANT distinction: 'producator' = the manufacturer of the product (e.g., Hebei Huayang Steel Pipe). 'companie' = the CERTIFICATION BODY that issued the certificate (e.g., TÜV Rheinland, TÜV SUD, Bureau Veritas, Lloyd's). These are DIFFERENT entities. Also extract material/product, expiration date, producer address. For 'nume_document', write 'Certificat CE PED' or 'Certificat CE'."
    },
    "CERTIFICAT_CALITATE": {
        "fields": ["material", "producator", "companie"],
        "instructions": "Extract: product/material name - the PHYSICAL PRODUCT being certified. Look for pipe type (PEHD, PVC, PE100), dimensions (DN, PN, SDR), and product description. Common materials: 'Teava PEHD PE100', 'Teava PVC multistrat', 'Fitinguri PEID'. If the product field is empty (template document) but a standard reference exists, DEDUCE the material from the standard: EN 12201 = 'Teava PE pentru apa (conform EN 12201)', EN 13476 = 'Teava PVC multistrat (conform EN 13476)', EN 1452 = 'Teava PVC-U presiune (conform EN 1452)', EN 10204 = 'Tevi/produse din otel (conform EN 10204)'. Also extract producer and company."
    },
    "CUI": {
        "fields": ["companie", "cui_number", "adresa_producator"],
        "instructions": "Extract: company name (EXACTLY as registered - common companies: TERAPLAST, VALROM INDUSTRIE, TEHNO WORLD, ZAKPREST CONSTRUCT), CUI number (just digits), registered address. For 'nume_document', use 'Certificat de Inregistrare'."
    },
    "ALTELE": {
        "fields": ["companie", "material"],
        "instructions": "Extract: any company name and material/product mentioned."
    },
}

EXTRACTION_SYSTEM_PROMPT = """You are a precise data extraction system for Romanian construction industry documents.
Extract ONLY the requested fields from the document text. Follow these rules strictly:

1. Return ONLY valid JSON with the exact field names requested
2. If a field is not found in the document, use null (not empty string)
3. For dates, use format DD.MM.YYYY when possible
4. For "data_expirare" (expiration date): look for "valabil pana la", "valid until", "data expirarii", "valabilitate", "expires", "date of expiry". If only issue date + validity period exist, CALCULATE the expiration date (e.g., issued 01.01.2024 + 3 years = 01.01.2027). ALWAYS try hard to find or calculate an expiration date.
5. For "material": extract the ACTUAL PRODUCT/MATERIAL name, NOT standard references.
   - CORRECT: "Tevi PEHD PE100 RC SDR11 DN110-DN630", "Fitinguri PVC-U PN10/PN16"
   - WRONG: "Executat dupa EN 12201-2" (this is a standard reference, not a material)
   - If the document is about a specific product, extract its commercial name and specifications
   - Include dimensions, standards references AFTER the product name if present
6. For "producator": the manufacturer/producer of the product.
   - Fix obvious OCR errors in company names (e.g., "TERAPIA" should be "TERAPLAST" if context shows Teraplast)
   - For well-known Romanian companies, use the correct name: TERAPLAST, VALROM, TEHNO WORLD
7. For "distribuitor": the company authorized to distribute the product
8. For "companie": if the company is neither clearly a producer nor distributor, use this field. For ISO certs, this is the certified company.
   - Same OCR correction rules as producator
9. For "adresa_producator": full address of the producer/manufacturer. Fix obvious OCR errors in addresses.
10. For "cui_number": just the numeric CUI code
11. For "standard_iso": the ISO standard number (e.g., "ISO 9001:2015", "ISO 14001:2015")
12. Fix obvious OCR errors in ALL extracted values:
    - "lndustrie" → "Industrie", "ser,;ce" → "Service", "TUV" → "TÜV"
    - "Să'ațel" → "Sărățel", "TERAPIA" → "TERAPLAST" (when context shows Teraplast)
    - Interpret Romanian diacritics: "Ţ" = "Ț", "ş" = "ș", "ã" = "ă"
13. For "nume_document": extract ONLY the document type/title, MAX 40 characters.
    - CORRECT examples: "Agrement Tehnic", "Aviz Sanitar", "Certificat CE PED", "Certificat ISO 9001", "Certificat de Inregistrare", "Fisa Tehnica Produs", "Declaratie de Performanta"
    - Do NOT include document numbers, dates, reference codes, or descriptions
    - Do NOT include newlines
    - For CE/PED certificates: use "Certificat CE PED"
    - For ISO certificates: use "Certificat ISO [standard]" (e.g., "Certificat ISO 9001")
    - WRONG: "Certificat Sistem de management al calității pentru Producător" (too long, includes description)

Respond with ONLY the JSON object, no markdown, no explanation."""


# ============================================================
# LEVEL 1: FILENAME REGEX CLASSIFICATION
# ============================================================

FILENAME_RULES = [
    # === HIGH PRIORITY: Very specific patterns first ===

    # Aviz Tehnic + Agrement (MUST be before individual patterns)
    (r'(?i)(aviz\s*(si|și|\+|&)\s*(at|agrement))|(aviz[- ]agrement)|(at\s*\+?\s*agt)', "AVIZ_TEHNIC_SI_AGREMENT", 0.95),

    # Aviz Sanitar - "sanitar" or "AVS" or "AS" (standalone) in filename
    (r'(?i)sanitar|\bAVS\b|(?:^|\b)\d+\.\s*AS\b', "AVIZ_SANITAR", 0.95),

    # CUI in filename
    (r'(?i)\bCUI\b', "CUI", 0.95),

    # ISO certificates - match "ISO" + optional standard number, or just "ISO" as standalone
    # Also match files with just standard numbers like "14001", "9001" that imply ISO
    (r'(?i)\bISO\s*\d{4,5}|\bISO\b', "ISO", 0.90),
    (r'(?i)\b(9001|14001|45001|50001)\b', "ISO", 0.80),

    # CE marking
    (r'(?i)(?:^|\b)CE(?:\s|[-_]|$)|(?i)marcaj\s*CE|(?i)CE\s*PED|(?i)excludere.*CE', "CE", 0.90),

    # Declaratie Performanta (before DC) - includes DP- prefix pattern
    (r'(?i)declara[tț]i[ea]\s*de\s*performan[tț][aă]|\bDOP\b|^DP[-\s]', "DECLARATIE_PERFORMANTA", 0.95),

    # Declaratie Conformitate - various patterns
    (r'(?i)declara[tț]i[ea]\s*de?\s*conformitate', "DECLARATIE_CONFORMITATE", 0.95),
    (r'(?i)(?:^|\b)\d+\.\s*DC\b', "DECLARATIE_CONFORMITATE", 0.90),
    (r'(?i)^DC\s+\d', "DECLARATIE_CONFORMITATE", 0.90),

    # Certificate de conformitate (variant naming)
    (r'(?i)certificate?\s*de\s*conformitate', "DECLARATIE_CONFORMITATE", 0.90),

    # Certificat Garantie
    (r'(?i)certificat\s*de?\s*garan[tț]ie|\bCG\b', "CERTIFICAT_GARANTIE", 0.90),

    # Certificat Calitate / Certificate 3.1 (EN 10204)
    (r'(?i)certificat\s*de?\s*calitate', "CERTIFICAT_CALITATE", 0.90),
    (r'(?i)(?:^|\b)\d+\.\s*CC\b', "CERTIFICAT_CALITATE", 0.85),
    (r'(?i)^CC\s+\d', "CERTIFICAT_CALITATE", 0.85),
    (r'(?i)(?:^|\b)\d+\.\s*C\s*3\.1\b', "CERTIFICAT_CALITATE", 0.93),
    (r'(?i)\bMTC\b|mill\s*test\s*cert', "CERTIFICAT_CALITATE", 0.93),

    # Agrement (after combined aviz+agrement)
    (r'(?i)\bagrement\b|\bAGT\b', "AGREMENT", 0.90),

    # Aviz Tehnic (after sanitar and combined)
    (r'(?i)\baviz\s*tehnic\b|\bAVT\b', "AVIZ_TEHNIC", 0.90),

    # Autorizatie
    (r'(?i)autorizat|authorization|autorizare', "AUTORIZATIE_DISTRIBUTIE", 0.90),

    # Fisa Tehnica - including "data sheet", catalog
    (r'(?i)fi[sș][aă]\s*tehnic[aă]|\bFT\b|data\s*sheet|technical\s*data', "FISA_TEHNICA", 0.90),
    (r'(?i)\bcatalog\b', "FISA_TEHNICA", 0.85),

    # Product-specific filenames that are typically FISA_TEHNICA
    # (Teava, Garnitura, Cot - product specifications)
    (r'(?i)^(teava|garnitura|cot|teu|reductie|mufa|dop|piesa)\s', "FISA_TEHNICA", 0.70),
]


def classify_by_filename(filename: str) -> tuple:
    """Level 1: Try to classify by filename patterns. Returns (category, confidence) or (None, 0)."""
    for pattern, category, conf in FILENAME_RULES:
        if re.search(pattern, filename):
            return category, conf
    return None, 0.0


# ============================================================
# PDF TEXT EXTRACTION (shared by classification + extraction)
# ============================================================

def extract_text_from_pdf(filepath: str, max_pages: int = 3) -> dict:
    """Extract text and metadata from PDF. Used by classification."""
    result = {"text": "", "page_count": 0, "has_tables": False, "extraction_method": "none"}

    try:
        with pdfplumber.open(filepath) as pdf:
            result["page_count"] = len(pdf.pages)
            texts = []
            for i, page in enumerate(pdf.pages[:max_pages]):
                text = page.extract_text() or ""
                texts.append(text)
                if not result["has_tables"] and page.extract_tables():
                    result["has_tables"] = True
            result["text"] = "\n".join(texts)
            if result["text"].strip():
                result["extraction_method"] = "pdfplumber"
    except Exception:
        pass

    if not result["text"].strip() or result["page_count"] == 0:
        try:
            doc = fitz.open(filepath)
            if result["page_count"] == 0:
                result["page_count"] = len(doc)
            if not result["text"].strip():
                texts = []
                for i, page in enumerate(doc):
                    if i >= max_pages:
                        break
                    texts.append(page.get_text())
                result["text"] = "\n".join(texts)
                if result["text"].strip():
                    result["extraction_method"] = "pymupdf"
            doc.close()
        except Exception:
            pass

    # OCR fallback for scanned documents
    if not result["text"].strip():
        try:
            doc = fitz.open(filepath)
            texts = []
            for i, page in enumerate(doc):
                if i >= max_pages:
                    break
                pix = page.get_pixmap(dpi=200)
                img = Image.open(io.BytesIO(pix.tobytes('png')))
                ocr_text = pytesseract.image_to_string(img, lang='eng')
                texts.append(ocr_text)
            result["text"] = "\n".join(texts)
            if result["text"].strip():
                result["extraction_method"] = "ocr"
            doc.close()
        except Exception:
            pass

    return result


def extract_full_text(filepath: str, max_pages: int = 5) -> str:
    """
    Extract maximum text from PDF for data extraction.
    Uses more pages and Romanian OCR for better extraction coverage.
    """
    text = ''

    # Try pdfplumber first (best for structured PDFs)
    try:
        with pdfplumber.open(filepath) as pdf:
            text = '\n'.join(p.extract_text() or '' for p in pdf.pages[:max_pages])
    except Exception:
        pass

    # Fallback: PyMuPDF
    if not text.strip():
        try:
            doc = fitz.open(filepath)
            text = '\n'.join(p.get_text() for p in doc[:max_pages])
            doc.close()
        except Exception:
            pass

    # OCR fallback with Romanian + English
    if not text.strip():
        try:
            doc = fitz.open(filepath)
            ocr_parts = []
            for i, page in enumerate(doc):
                if i >= max_pages:
                    break
                pix = page.get_pixmap(dpi=250)
                img = Image.open(io.BytesIO(pix.tobytes('png')))
                # Try Romanian+English first, fall back to just English
                try:
                    ocr_text = pytesseract.image_to_string(img, lang='ron+eng')
                except Exception:
                    try:
                        ocr_text = pytesseract.image_to_string(img, lang='eng')
                    except Exception:
                        ocr_text = ''
                ocr_parts.append(ocr_text)
            text = '\n'.join(ocr_parts)
            doc.close()
        except Exception:
            pass

    return text.strip()


# ============================================================
# COMPANY NAME NORMALIZATION
# ============================================================

def normalize_company_name(name: str) -> str:
    """
    Normalize Romanian company names to a consistent format.
    Removes SC/S.C. prefix, normalizes SRL/S.R.L./SA/S.A. suffix,
    normalizes capitalization, and standardizes known companies.
    """
    if not name:
        return name

    val = name.strip()

    # Step 1: Remove SC / S.C. prefix (archaic, no longer required)
    val = re.sub(r'^S\.?\s*C\.?\s+', '', val, flags=re.IGNORECASE).strip()

    # Step 2: Normalize the legal form suffix
    # Match S.R.L. / S.R.L / SRL / srl at end of string
    val = re.sub(r'\s+S\.?\s*R\.?\s*L\.?\s*$', ' SRL', val, flags=re.IGNORECASE)
    # Match S.A. / SA / S.A at end of string
    val = re.sub(r'\s+S\.?\s*A\.?\s*$', ' SA', val, flags=re.IGNORECASE)
    # Also handle hyphenated form like "TERAPLAST-SA"
    val = re.sub(r'-S\.?\s*A\.?\s*$', ' SA', val, flags=re.IGNORECASE)

    # Step 3: Known company name normalization
    # Map common variations to canonical names
    KNOWN_COMPANIES = {
        "TERAPLAST": "TERAPLAST SA",
        "TERAPLAST SA": "TERAPLAST SA",
        "TERAPLAST SRL": "TERAPLAST SA",
        "TERAPIA": "TERAPLAST SA",
        "VALROM INDUSTRIE": "VALROM INDUSTRIE SRL",
        "VALROM INDUSTRIE SRL": "VALROM INDUSTRIE SRL",
        "VALROM": "VALROM INDUSTRIE SRL",
        "VALROM INDUSTRIE SA": "VALROM INDUSTRIE SRL",
        "TEHNO WORLD": "TEHNO WORLD SRL",
        "TEHNO WORLD SRL": "TEHNO WORLD SRL",
        "TEHNOWORLD": "TEHNO WORLD SRL",
        "TEHNOWORLD SRL": "TEHNO WORLD SRL",
        "ZAKPREST CONSTRUCT": "ZAKPREST CONSTRUCT SRL",
        "ZAKPREST CONSTRUCT SRL": "ZAKPREST CONSTRUCT SRL",
    }

    # Try to match the cleaned name (uppercase) against known companies
    upper_val = val.upper().strip()
    if upper_val in KNOWN_COMPANIES:
        return KNOWN_COMPANIES[upper_val]

    return val.strip()


# ============================================================
# DATA EXTRACTION VIA AI
# ============================================================

def extract_data_with_ai(text: str, category: str, filename: str,
                          model: str = None, retries: int = 2) -> dict:
    """
    Extract structured data from document text using AI.
    Returns dict with extracted fields or {"error": "..."} on failure.
    """
    if model is None:
        model = EXTRACTION_MODEL

    schema = EXTRACTION_SCHEMA.get(category, EXTRACTION_SCHEMA["ALTELE"])

    # Build the universal output fields list
    output_fields = list(set(schema["fields"] + ["nume_document"]))

    user_prompt = f"""Document filename: {filename}
Document category: {category}
Required fields: {json.dumps(output_fields)}

Instructions: {schema['instructions']}

Also extract:
- "nume_document": the official title of the document as it appears in the header/title area

Document text:
---
{text[:5000]}
---

Return a JSON object with the requested fields. Use null for fields not found in the document."""

    last_error = None
    for attempt in range(retries + 1):
        try:
            resp = requests.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {OPENROUTER_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": model,
                    "messages": [
                        {"role": "system", "content": EXTRACTION_SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                    "temperature": 0.0,
                    "max_tokens": 1000,
                },
                timeout=45,
            )
            resp.raise_for_status()
            content = resp.json()["choices"][0]["message"]["content"]
            # Clean markdown if present
            content = re.sub(r'^```json\s*', '', content.strip())
            content = re.sub(r'\s*```$', '', content.strip())
            return json.loads(content)
        except json.JSONDecodeError as e:
            last_error = f"JSON parse error: {e}"
            # Try to extract JSON from mixed content
            try:
                match = re.search(r'\{[^{}]*\}', content, re.DOTALL)
                if match:
                    return json.loads(match.group())
            except Exception:
                pass
        except requests.exceptions.Timeout:
            last_error = "Request timeout"
            time.sleep(2)
        except Exception as e:
            last_error = str(e)
            if attempt < retries:
                time.sleep(1)

    return {"error": last_error or "Unknown error"}


def normalize_extraction_result(raw: dict, category: str) -> dict:
    """
    Normalize AI extraction result to standard output columns:
    Nume Document, Material, Data Expirare, Companie, Producator, Distribuitor, Adresa Producator
    """
    result = {
        "nume_document": None,
        "material": None,
        "data_expirare": None,
        "companie": None,
        "producator": None,
        "distribuitor": None,
        "adresa_producator": None,
    }

    if "error" in raw:
        return result

    # Direct mappings
    for key in result:
        if key in raw and raw[key]:
            result[key] = str(raw[key]).strip()

    # Normalize company names in all company fields
    for company_field in ("companie", "producator", "distribuitor"):
        if result.get(company_field):
            result[company_field] = normalize_company_name(result[company_field])

    # Handle "data_emitere" + "valabilitate" → "data_expirare" calculation
    if not result["data_expirare"] and raw.get("data_emitere") and raw.get("valabilitate"):
        valab = str(raw["valabilitate"]).lower()
        if "contract" in valab or "nedeterminat" in valab:
            result["data_expirare"] = "Pe durata contractului"
        else:
            result["data_expirare"] = f"Emitere: {raw['data_emitere']}, Valabilitate: {raw['valabilitate']}"

    # For ISO certs, the "companie" is the certified company
    if category == "ISO" and raw.get("companie") and not result["companie"]:
        result["companie"] = str(raw["companie"]).strip()

    # For CUI, map cui_number and adresa
    if category == "CUI":
        if raw.get("cui_number"):
            # Store CUI number in companie field with prefix
            if result["companie"]:
                result["companie"] = f"{result['companie']} (CUI: {raw['cui_number']})"
            else:
                result["companie"] = f"CUI: {raw['cui_number']}"
        if raw.get("adresa") and not result["adresa_producator"]:
            result["adresa_producator"] = str(raw["adresa"]).strip()

    # For ISO: standard goes in nume_document (e.g., "Certificat ISO 9001"), NOT in material
    # ISO certificates certify management systems, not physical materials
    if category == "ISO":
        result["material"] = None  # ISO is not a material
        if raw.get("standard_iso") and result.get("nume_document"):
            std = str(raw["standard_iso"]).strip()
            # Ensure nume_document includes the standard
            if "iso" not in (result["nume_document"] or "").lower():
                result["nume_document"] = f"Certificat {std}"

    # Clean up values: remove newlines, trim, handle null-like values, fix OCR
    for key in result:
        if result[key]:
            val = str(result[key]).strip()

            # Remove newlines and excessive whitespace
            val = re.sub(r'\s*\n\s*', ' ', val).strip()
            # Remove double/trailing quotes artifacts
            val = re.sub(r'"+$', '', val).strip()
            val = re.sub(r'^"+', '', val).strip()

            # Clean and truncate nume_document
            if key == "nume_document":
                val = re.sub(r'\s+\d{3}[-/]\d{2}[-/]\d+[-/]?\d*$', '', val).strip()
                val = re.sub(r'\s+Nr\.?\s*\d+.*$', '', val).strip()
                # Truncate to max 50 chars at a word boundary
                if len(val) > 50:
                    cut = val[:50].rfind(' ')
                    if cut > 20:
                        val = val[:cut].rstrip(' ,;')
                    else:
                        val = val[:50]

            # Fix common OCR errors everywhere
            ocr_fixes = [
                ("lndustrie", "Industrie"),
                ("ser,;ce", "Service"),
                ("ser;ce", "Service"),
                ("lron", "Iron"),
                ("PEÎD", "PEID"),
                ("PEÎ D", "PEID"),
            ]
            for old_ocr, new_ocr in ocr_fixes:
                val = val.replace(old_ocr, new_ocr)

            # Normalize Romanian diacritics (cedilla → comma below)
            diacritics_map = {
                'Ţ': 'Ț', 'ţ': 'ț',
                'Ş': 'Ș', 'ş': 'ș',
                'ã': 'ă',
            }
            for old_d, new_d in diacritics_map.items():
                val = val.replace(old_d, new_d)

            # Truncate excessively long material descriptions (max 300 chars)
            if key == "material" and len(val) > 300:
                cut_pos = val[:300].rfind(',')
                if cut_pos < 150:
                    cut_pos = val[:300].rfind(';')
                if cut_pos < 150:
                    cut_pos = 297
                val = val[:cut_pos].rstrip(' ,;') + "..."

            # Truncate excessively long addresses (max 250 chars)
            if key == "adresa_producator" and len(val) > 250:
                val = val[:250].rstrip(' ,;')

            # Check for null-like values
            if val.lower() in ('null', 'none', 'n/a', '-', ''):
                val = None

            result[key] = val

    return result


def extract_document_data(filepath: str, category: str, model: str = None,
                           preloaded_text: str = None) -> dict:
    """
    Full extraction pipeline for a single document.
    Returns normalized dict with standard columns.
    """
    filename = os.path.basename(filepath)

    # Get text (use preloaded if available, otherwise extract fresh with more pages)
    if preloaded_text and len(preloaded_text.strip()) > 50:
        text = preloaded_text
    else:
        text = extract_full_text(filepath, max_pages=5)

    if not text or len(text.strip()) < 20:
        return {
            "nume_document": None, "material": None, "data_expirare": None,
            "companie": None, "producator": None, "distribuitor": None,
            "adresa_producator": None, "extraction_status": "NO_TEXT",
        }

    # AI extraction
    raw = extract_data_with_ai(text, category, filename, model=model)

    if "error" in raw:
        return {
            "nume_document": None, "material": None, "data_expirare": None,
            "companie": None, "producator": None, "distribuitor": None,
            "adresa_producator": None, "extraction_status": f"ERROR: {raw['error'][:80]}",
        }

    # Normalize to standard columns
    result = normalize_extraction_result(raw, category)
    result["extraction_status"] = "OK"
    return result


# ============================================================
# LEVEL 2: TEXT-BASED CLASSIFICATION (Rule-based, no AI)
# ============================================================

def classify_by_text(text: str, page_count: int, has_tables: bool) -> tuple:
    """Level 2: Classify based on extracted text content. Returns (category, confidence, reasoning)."""
    if not text or len(text.strip()) < 20:
        return None, 0.0, "No text available"

    text_lower = text.lower()
    first_500 = text_lower[:500]

    # ---- AVIZ SANITAR: "aviz sanitar" as document TITLE (not just mentioned) ----
    if re.search(r'aviz\s+sanitar', first_500):
        return "AVIZ_SANITAR", 0.95, "Title contains 'Aviz Sanitar'"

    # ---- WRAS APPROVAL (before CE/ISO, because WRAS docs reference ISO/CE in passing) ----
    if re.search(r'\bwras\b', text_lower) and re.search(r'approv|aproba', text_lower):
        return "CERTIFICAT_CALITATE", 0.92, "WRAS material approval certificate"

    # ---- CE MARKING (check BEFORE DC because "EU Certificate of Conformity" for PED = CE) ----
    ce_markers = [
        r'directiva\s+\d{4}/\d+',
        r'directive\s+\d{4}/\d+',
        r'(?<!/)\bce\s+mark(?!\.)',
        r'ped\s+certificate',
        r'pressure\s+equipment\s+directive',
        r'pressure\s+equipment\b',
        r'european\s+conformity',
        r'eu\s+certificate\s+of\s+conformity',
        r'annex\s+iii.*module\s+h',
    ]
    ce_score = sum(1 for p in ce_markers if re.search(p, text_lower))
    if ce_score >= 2:
        return "CE", 0.92, f"CE directive/marking references found ({ce_score} markers)"
    if ce_score == 1 and re.search(r'ce\s+mark|ped\s+certificate|eu\s+certificate|european\s+conformity', text_lower):
        return "CE", 0.88, "Strong CE marker found"

    # ---- DECLARATIE DE PERFORMANTA (before DC, since DC markers are broader) ----
    dp_markers = [
        r'declara[tț]i[ea]\s+de\s+performan[tț][aă]',
        r'declaration\s+of\s+performance',
        r'\bdop\s+nr',
        r'regulamentul.*305/2011',
        r'regulation.*305/2011',
        r'cod\s+unic\s+de\s+identificare\s+al\s+produsului',
    ]
    dp_score = sum(1 for p in dp_markers if re.search(p, text_lower))
    dp_in_title = sum(1 for p in dp_markers if re.search(p, first_500))
    if dp_score >= 2 and dp_in_title >= 1:
        return "DECLARATIE_PERFORMANTA", 0.92, f"Text contains {dp_score} DOP markers ({dp_in_title} in title)"
    if dp_in_title >= 1 and re.search(r'declara[tț]i[ea]\s+de\s+performan', first_500):
        return "DECLARATIE_PERFORMANTA", 0.88, "Title contains 'Declaratie de Performanta'"

    # ---- DECLARATIE DE CONFORMITATE ----
    dc_markers = [
        r'declara[tț]i[ea]\s+de\s+conformitate',
        r'certificat\s+de\s+conformitate',
        r'hg\s*(?:nr\.?)?\s*668',
        r'hot[aă]r[aâ]rea\s+guvernului',
        r'asigur[aă]m.*garantăm.*declar[aă]m',
        r'nu\s+pun\s+[îi]n\s+pericol\s+via[tț]a',
        r'declara[tț]i[ea]\s+nr\.',
        r'certificate?\s+of\s+conformity(?!.*pressure)',
    ]
    dc_score = sum(1 for p in dc_markers if re.search(p, text_lower))
    if dc_score >= 2:
        return "DECLARATIE_CONFORMITATE", 0.90, f"Text contains {dc_score} DC markers"
    if dc_score == 1 and re.search(r'declara[tț]i[ea]\s+de\s+conformitate', first_500):
        return "DECLARATIE_CONFORMITATE", 0.85, "Title contains 'Declaratie de Conformitate'"
    if dc_score == 1 and re.search(r'certificat\s+de\s+conformitate', first_500):
        return "DECLARATIE_CONFORMITATE", 0.85, "Title contains 'Certificat de Conformitate'"

    # ---- ISO MANAGEMENT SYSTEM CERTIFICATES ----
    iso_mgmt_title = re.search(r'iso\s*(9001|14001|45001|50001)', first_500)
    iso_mgmt_body = re.search(r'iso\s*(9001|14001|45001|50001)', text_lower)
    iso_mgmt = iso_mgmt_title or iso_mgmt_body
    if iso_mgmt:
        cert_context = any(w in first_500 for w in [
            'certificate', 'certificat', 'certification', 'certified',
            'scope', 'domeniu', 'validity', 'valabil', 'accredit',
            'management system', 'sistem de management'
        ])
        if cert_context and iso_mgmt_title:
            return "ISO", 0.95, f"ISO {iso_mgmt.group(1)} certificate in title"
        if cert_context and not iso_mgmt_title:
            return "ISO", 0.78, f"ISO {iso_mgmt.group(1)} in body with cert context"
        if iso_mgmt_title:
            return "ISO", 0.80, f"ISO {iso_mgmt.group(1)} in title without cert context"

    # ---- AGREMENT TEHNIC ----
    agrement_markers = [
        r'agrement\s+tehnic',
        r'ministerul\s+dezvolt[aă]rii',
        r'consiliul\s+tehnic\s+permanent',
        r'grupa\s+specializat[aă]',
    ]
    agr_score = sum(1 for p in agrement_markers if re.search(p, text_lower))
    if agr_score >= 2:
        if re.search(r'aviz\s+tehnic', first_500):
            return "AVIZ_TEHNIC_SI_AGREMENT", 0.90, "Both Aviz Tehnic and Agrement in text"
        return "AGREMENT", 0.90, f"Text contains {agr_score} Agrement markers"

    # ---- AVIZ TEHNIC ----
    if re.search(r'aviz\s+tehnic', first_500) and not re.search(r'agrement', first_500):
        return "AVIZ_TEHNIC", 0.85, "Title contains 'Aviz Tehnic' without Agrement"

    # ---- CERTIFICAT CALITATE ----
    if re.search(r'certificat\s+de\s+calitate', first_500):
        return "CERTIFICAT_CALITATE", 0.90, "Title: Certificat de Calitate"

    # ---- CERTIFICAT GARANTIE ----
    if re.search(r'certificat\s+de\s+garan[tț]ie', first_500):
        return "CERTIFICAT_GARANTIE", 0.90, "Title: Certificat de Garantie"

    # ---- FISA TEHNICA (title check - before AUTORIZATIE and CUI) ----
    if re.search(r'fi[sș][aă]\s+tehnic[aă]', first_500):
        return "FISA_TEHNICA", 0.90, "Title: Fisa Tehnica"
    if re.search(r'technical\s+data\s+sheet', first_500):
        return "FISA_TEHNICA", 0.90, "Title: Technical Data Sheet"

    # ---- AUTORIZATIE DISTRIBUTIE (before CUI) ----
    autorizatie_markers = [
        r'autoriz[aă]m\s+prin\s+prezenta',
        r's[aă]\s+distribui[ea]',
        r'autorizat\s+s[aă]',
        r'authorization\s+letter',
        r'pe\s+teritoriul\s+rom[aâ]niei',
        r'autorizare\s+de\s+comercializare',
        r'authorized.*distribut',
    ]
    aut_score = sum(1 for p in autorizatie_markers if re.search(p, text_lower))
    if aut_score >= 2:
        return "AUTORIZATIE_DISTRIBUTIE", 0.90, f"Text contains {aut_score} authorization markers"
    if aut_score == 1 and re.search(r'autorizare\s+de\s+comercializare', first_500):
        return "AUTORIZATIE_DISTRIBUTIE", 0.85, "Title contains 'Autorizare de Comercializare'"

    # ---- CUI ----
    cui_indicators = [
        re.search(r'certificat\s+de\s+[îi]nregistrare', text_lower),
        re.search(r'cod\s+unic\s+de\s+[îi]nregistrare', text_lower),
        re.search(r'registrul\s+comer[tț]ului', text_lower),
        re.search(r'oficiul\s+registrului', text_lower),
    ]
    if sum(1 for x in cui_indicators if x) >= 2:
        if not re.search(r'declara[tț]i[ea]\s+de\s+performan', text_lower) and \
           not re.search(r'fi[sș][aă]\s+tehnic[aă]|domeniu\s+de\s+utilizare|specifica[tț]i', text_lower):
            return "CUI", 0.90, "Company registration document (2+ indicators)"
    if sum(1 for x in cui_indicators if x) >= 1:
        if re.search(r'certificat\s+de\s+[îi]nregistrare', first_500):
            return "CUI", 0.88, "Title: Certificat de Inregistrare"

    # ---- DECLARATIE PERFORMANTA ----
    if re.search(r'declara[tț]i[ea]\s+de\s+performan[tț][aă]', text_lower) or \
       re.search(r'regulamentul.*305/2011|regulation.*305/2011', text_lower):
        return "DECLARATIE_PERFORMANTA", 0.90, "Declaration of Performance markers found"

    # ---- FISA TEHNICA ----
    ft_markers = [
        r'fi[sș][aă]\s+tehnic[aă]',
        r'technical\s+data\s+sheet',
        r'domeniu\s+de\s+utilizare',
        r'norma\s+de\s+produs',
        r'specifica[tț]i[ei]\s+tehnic[ei]',
    ]
    ft_score = sum(1 for p in ft_markers if re.search(p, text_lower))

    product_spec_markers = [
        r'\bDN\s*\d+',
        r'\bPN\s*\d+',
        r'\bSDR\s*\d+',
        r'\bDE\s*\d+',
        r'dimensiun[ie]',
        r'greutate',
        r'material[:\s]',
        r'diametr',
        r'grosime.*pere[tț]',
        r'\bmm\b.*\bmm\b',
    ]
    spec_score = sum(1 for p in product_spec_markers if re.search(p, text_lower))

    if ft_score >= 1:
        return "FISA_TEHNICA", 0.90, f"Technical data sheet markers ({ft_score})"
    if spec_score >= 3 and page_count >= 1:
        return "FISA_TEHNICA", 0.80, f"Product specification with {spec_score} technical markers"
    if spec_score >= 2 and has_tables:
        return "FISA_TEHNICA", 0.75, f"Has tables + {spec_score} technical spec markers"

    return None, 0.0, "No confident text match"


# ============================================================
# LEVEL 3: AI CLASSIFICATION VIA OPENROUTER
# ============================================================

CLASSIFICATION_PROMPT = """Classify this Romanian construction document into exactly ONE category.

CATEGORIES: AGREMENT, AUTORIZATIE_DISTRIBUTIE, AVIZ_TEHNIC_SI_AGREMENT, AVIZ_TEHNIC, AVIZ_SANITAR, CE, CERTIFICAT_CALITATE, CERTIFICAT_GARANTIE, CUI, DECLARATIE_CONFORMITATE, DECLARATIE_PERFORMANTA, FISA_TEHNICA, ISO, ALTELE

DOCUMENT:
Filename: {filename}
Pages: {page_count}
Text: {text}

Respond with ONLY valid JSON: {{"category": "NAME", "confidence": 0.95, "reasoning": "brief"}}"""


def classify_by_ai(filename: str, pdf_info: dict) -> tuple:
    """Level 3: AI classification via OpenRouter."""
    text_preview = pdf_info["text"][:2000] if pdf_info["text"] else "(scanned/no text)"
    prompt = CLASSIFICATION_PROMPT.format(
        filename=filename, page_count=pdf_info["page_count"], text=text_preview
    )

    VALID_CATEGORIES = {
        "AGREMENT", "AUTORIZATIE_DISTRIBUTIE", "AVIZ_TEHNIC_SI_AGREMENT",
        "AVIZ_TEHNIC", "AVIZ_SANITAR", "CE", "CERTIFICAT_CALITATE",
        "CERTIFICAT_GARANTIE", "CUI", "DECLARATIE_CONFORMITATE",
        "DECLARATIE_PERFORMANTA", "FISA_TEHNICA", "ISO", "ALTELE",
    }

    try:
        response = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENROUTER_API_KEY}", "Content-Type": "application/json"},
            json={"model": AI_MODEL, "messages": [{"role": "user", "content": prompt}],
                  "temperature": 0.1, "max_tokens": 200},
            timeout=30
        )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"].strip()
        content = re.sub(r'^```json\s*', '', content)
        content = re.sub(r'\s*```$', '', content)
        result = json.loads(content)
        category = result.get("category", "ALTELE")
        if category not in VALID_CATEGORIES:
            category = "ALTELE"
        return category, result.get("confidence", 0.5), result.get("reasoning", "")
    except Exception as e:
        return None, 0.0, f"AI unavailable: {e}"


# ============================================================
# COMBINED CLASSIFIER
# ============================================================

def classify_document(filepath: str, use_ai: bool = True) -> dict:
    """Classify a single document using 3-level approach."""
    filename = os.path.basename(filepath)

    # Level 1: Filename regex
    fn_category, fn_confidence = classify_by_filename(filename)

    # If filename gives high confidence (>= 0.90), trust it directly
    if fn_category and fn_confidence >= 0.90:
        return {
            "filepath": filepath,
            "filename": filename,
            "predicted_category": fn_category,
            "confidence": fn_confidence,
            "method": "filename_regex",
            "reasoning": f"Filename clearly matches {fn_category}",
        }

    # Level 2: Extract text and analyze
    pdf_info = extract_text_from_pdf(filepath)
    text_category, text_confidence, text_reasoning = classify_by_text(
        pdf_info["text"], pdf_info["page_count"], pdf_info["has_tables"]
    )

    # If filename gave a medium-confidence match, cross-check with text
    if fn_category and fn_confidence >= 0.70:
        if text_category and text_category == fn_category:
            return {
                "filepath": filepath,
                "filename": filename,
                "predicted_category": fn_category,
                "confidence": min(0.99, max(fn_confidence, text_confidence) + 0.05),
                "method": "filename+text_agree",
                "reasoning": f"Filename and text both indicate {fn_category}",
            }
        elif text_category and text_confidence > fn_confidence:
            return {
                "filepath": filepath,
                "filename": filename,
                "predicted_category": text_category,
                "confidence": text_confidence,
                "method": "text_override",
                "reasoning": f"Text ({text_confidence:.0%}) overrides filename ({fn_confidence:.0%}): {text_reasoning}",
            }
        else:
            return {
                "filepath": filepath,
                "filename": filename,
                "predicted_category": fn_category,
                "confidence": fn_confidence,
                "method": "filename_regex",
                "reasoning": f"Filename matches {fn_category} (text inconclusive)",
            }

    # Text-only classification (filename didn't match)
    if text_category and text_confidence >= 0.75:
        return {
            "filepath": filepath,
            "filename": filename,
            "predicted_category": text_category,
            "confidence": text_confidence,
            "method": "text_rules",
            "reasoning": text_reasoning,
        }

    # Level 3: AI fallback
    if use_ai:
        ai_cat, ai_conf, ai_reason = classify_by_ai(filename, pdf_info)
        if ai_cat and ai_conf > 0:
            return {
                "filepath": filepath,
                "filename": filename,
                "predicted_category": ai_cat,
                "confidence": ai_conf,
                "method": "ai",
                "reasoning": ai_reason,
            }

    # Final fallback
    if fn_category:
        return {
            "filepath": filepath,
            "filename": filename,
            "predicted_category": fn_category,
            "confidence": fn_confidence,
            "method": "filename_low_conf",
            "reasoning": f"Low confidence filename match: {fn_category}",
        }
    if text_category:
        return {
            "filepath": filepath,
            "filename": filename,
            "predicted_category": text_category,
            "confidence": text_confidence,
            "method": "text_low_conf",
            "reasoning": text_reasoning,
        }

    return {
        "filepath": filepath,
        "filename": filename,
        "predicted_category": "ALTELE",
        "confidence": 0.1,
        "method": "fallback",
        "reasoning": "No classification method succeeded",
    }


# ============================================================
# TESTING & EVALUATION
# ============================================================

def load_ground_truth(base_dir: str) -> list:
    documents = []
    for folder_name, category in FOLDER_TO_CATEGORY.items():
        folder_path = os.path.join(base_dir, folder_name)
        if not os.path.exists(folder_path):
            continue
        for f in sorted(os.listdir(folder_path)):
            if f.lower().endswith('.pdf'):
                documents.append({
                    "filepath": os.path.join(folder_path, f),
                    "filename": f,
                    "true_category": category,
                })
    return documents


def run_evaluation(use_ai: bool = False, save_results: str = None):
    documents = load_ground_truth(BASE_DIR)
    print(f"\n{'='*80}")
    print(f"CLASSIFICATION v2 EVALUATION")
    print(f"{'='*80}")
    print(f"Documents: {len(documents)} | AI: {'enabled' if use_ai else 'disabled'} | Model: {AI_MODEL}")
    print()

    results = []
    correct = 0
    total = 0
    errors = []
    category_stats = defaultdict(lambda: {"correct": 0, "total": 0})
    methods = defaultdict(int)

    for i, doc in enumerate(documents):
        total += 1
        result = classify_document(doc["filepath"], use_ai=use_ai)
        result["true_category"] = doc["true_category"]
        result["is_correct"] = result["predicted_category"] == doc["true_category"]
        results.append(result)
        methods[result["method"]] += 1

        true_cat = doc["true_category"]
        pred_cat = result["predicted_category"]
        category_stats[true_cat]["total"] += 1

        if result["is_correct"]:
            correct += 1
            category_stats[true_cat]["correct"] += 1
            status = "✓"
        else:
            status = "✗"
            errors.append(result)

        print(f"  [{i+1:3d}/{len(documents)}] {status} {doc['filename'][:50]:50s} "
              f"TRUE={true_cat:30s} PRED={pred_cat:30s} [{result['method']}]")

    accuracy = correct / total * 100 if total else 0
    print(f"\n{'='*80}")
    print(f"RESULTS: {correct}/{total} = {accuracy:.1f}%")
    print(f"{'='*80}")

    print(f"\nMethods used:")
    for method, count in sorted(methods.items(), key=lambda x: -x[1]):
        print(f"  {method:25s}: {count:3d} ({count/total*100:.0f}%)")

    print(f"\n{'Category':<35s} {'Score':>10s} {'Acc':>7s}")
    print("-" * 55)
    for cat in sorted(category_stats.keys()):
        s = category_stats[cat]
        acc = s["correct"] / s["total"] * 100 if s["total"] else 0
        mark = "✓" if acc == 100 else "⚠️"
        print(f"  {cat:<33s} {s['correct']:>3d}/{s['total']:<3d}   {acc:>5.1f}% {mark}")

    if errors:
        print(f"\n{'='*80}")
        print(f"ERRORS ({len(errors)}):")
        print(f"{'='*80}")
        for e in errors:
            print(f"  {e['filename'][:55]:55s}")
            print(f"    TRUE: {e['true_category']:30s} PRED: {e['predicted_category']}")
            print(f"    Method: {e['method']} | Reason: {e['reasoning'][:80]}")
            print()

    if save_results:
        with open(save_results, 'w', encoding='utf-8') as f:
            json.dump({"accuracy": accuracy, "correct": correct, "total": total,
                       "errors": [{"filename": e["filename"], "true": e["true_category"],
                                   "predicted": e["predicted_category"], "method": e["method"],
                                   "reasoning": e["reasoning"]} for e in errors],
                       "results": results}, f, ensure_ascii=False, indent=2)
        print(f"Results saved to: {save_results}")

    return results, accuracy


# ============================================================
# PRODUCTION: CLASSIFY + EXTRACT
# ============================================================

def classify_folder(input_dir: str, output_dir: str = None, move_files: bool = False,
                    report_path: str = None, do_extract: bool = True,
                    extract_model: str = None):
    """
    PRODUCTION MODE: Classify + optionally extract data from all PDFs in a folder.

    Args:
        input_dir: Folder containing PDFs to classify
        output_dir: Where to create category subfolders (if move_files=True)
        move_files: Whether to copy files into category folders
        report_path: Path for Excel/CSV report
        do_extract: Whether to run data extraction (default True)
        extract_model: Model to use for extraction (default EXTRACTION_MODEL)
    """
    import shutil

    # Find all PDFs
    pdfs = []
    for root, dirs, files in os.walk(input_dir):
        for f in sorted(files):
            if f.lower().endswith('.pdf'):
                pdfs.append(os.path.join(root, f))

    print(f"\nClassifying {len(pdfs)} PDFs from: {input_dir}")
    if do_extract:
        print(f"Extraction model: {extract_model or EXTRACTION_MODEL}")
    print(f"{'='*70}")

    results = []
    category_counts = defaultdict(int)
    extraction_results = []
    extract_errors = 0
    extract_ok = 0

    for i, filepath in enumerate(pdfs):
        # Step 1: Classify
        result = classify_document(filepath, use_ai=True)
        results.append(result)
        cat = result["predicted_category"]
        category_counts[cat] += 1

        conf_str = f"{result['confidence']:.0%}"
        print(f"  [{i+1:3d}/{len(pdfs)}] {result['filename'][:50]:50s} → {cat:25s} ({conf_str}, {result['method']})", end="")

        # Step 2: Extract data
        if do_extract:
            ext_data = extract_document_data(filepath, cat, model=extract_model)
            extraction_results.append(ext_data)

            # Get page count
            try:
                with pdfplumber.open(filepath) as pdf:
                    ext_data["numar_pagini"] = len(pdf.pages)
            except Exception:
                try:
                    doc = fitz.open(filepath)
                    ext_data["numar_pagini"] = len(doc)
                    doc.close()
                except Exception:
                    ext_data["numar_pagini"] = None

            status = ext_data.get("extraction_status", "?")
            if status == "OK":
                extract_ok += 1
                # Show key extracted fields
                mat = ext_data.get("material", "")
                prod = ext_data.get("producator", "")
                brief = f" | {prod[:25]}" if prod else ""
                brief += f" | {mat[:30]}" if mat else ""
                print(f"  ✓{brief}")
            else:
                extract_errors += 1
                print(f"  ✗ {status[:40]}")
        else:
            extraction_results.append({})
            print()

        # Move/copy file if requested
        if move_files and output_dir:
            CAT_TO_FOLDER = {
                "AGREMENT": "Agrement",
                "AUTORIZATIE_DISTRIBUTIE": "Autorizatii de comercializare sau de distributie",
                "AVIZ_TEHNIC_SI_AGREMENT": "Aviz Tehnic + Agrement",
                "AVIZ_TEHNIC": "Aviz tehnic",
                "AVIZ_SANITAR": "Avize Sanitare",
                "CE": "CE",
                "CERTIFICAT_CALITATE": "Certificat Calitate",
                "CERTIFICAT_GARANTIE": "Certificat Garantie",
                "CUI": "Cui",
                "DECLARATIE_CONFORMITATE": "Declaratii conformitate",
                "DECLARATIE_PERFORMANTA": "Declaratie de Performanta",
                "FISA_TEHNICA": "Fisa Tehnica",
                "ISO": "Iso",
                "ALTELE": "Z.Altele",
            }
            dest_folder = os.path.join(output_dir, CAT_TO_FOLDER.get(cat, "Z.Altele"))
            os.makedirs(dest_folder, exist_ok=True)
            dest_path = os.path.join(dest_folder, result["filename"])
            if not os.path.exists(dest_path):
                shutil.copy2(filepath, dest_path)

    # Summary with review flags
    needs_check = []
    needs_review = []
    for r in results:
        conf = r["confidence"]
        method = r["method"]
        if method == "ai" or method == "text_override":
            needs_check.append(r)
        elif conf < 0.85 or (method == "text_rules" and conf < 0.90):
            needs_review.append(r)

    print(f"\n{'='*70}")
    print(f"CLASSIFICATION COMPLETE: {len(pdfs)} documents")
    print(f"{'='*70}")

    for cat in sorted(category_counts.keys()):
        print(f"  {cat:35s}: {category_counts[cat]:3d}")

    # Method breakdown
    method_counts = defaultdict(int)
    for r in results:
        method_counts[r["method"]] += 1
    print(f"\nMetode utilizate:")
    for m, c in sorted(method_counts.items(), key=lambda x: -x[1]):
        print(f"  {m:25s}: {c:3d} ({c/len(results)*100:.0f}%)")

    # Review summary
    ok_count = len(results) - len(needs_check) - len(needs_review)
    print(f"\nReview status:")
    print(f"  OK (confident)          : {ok_count:3d} ({ok_count/len(results)*100:.0f}%)")
    print(f"  REVIEW (low confidence) : {len(needs_review):3d} ({len(needs_review)/len(results)*100:.0f}%)")
    print(f"  NEEDS CHECK (AI/override): {len(needs_check):3d} ({len(needs_check)/len(results)*100:.0f}%)")

    if do_extract:
        print(f"\nExtraction status:")
        print(f"  OK:     {extract_ok:3d} ({extract_ok/len(results)*100:.0f}%)")
        print(f"  Errors: {extract_errors:3d} ({extract_errors/len(results)*100:.0f}%)")

    if needs_check:
        print(f"\n  Documente care necesita verificare manuala:")
        for r in needs_check:
            print(f"    {r['filename'][:55]:55s} → {r['predicted_category']:25s} ({r['confidence']:.0%}, {r['method']})")

    # Generate report
    if report_path:
        generate_excel_report(results, extraction_results, report_path,
                              category_counts, method_counts, needs_check, needs_review,
                              do_extract)

    return results, extraction_results


def generate_excel_report(results, extraction_results, report_path,
                          category_counts, method_counts, needs_check, needs_review,
                          has_extraction=True):
    """Generate comprehensive Excel report with classification + extraction data."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Raport Documente"

    # === HEADERS ===
    if has_extraction:
        headers = [
            "Nr.", "Fisier", "Categorie", "Confidence", "Metoda",
            "Nume Document", "Numar Pagini", "Material", "Data Expirare",
            "Companie", "Producator", "Distribuitor", "Adresa Producator",
            "Review", "Corectie Manuala", "Status Extractie", "Reasoning", "Cale Fisier"
        ]
    else:
        headers = [
            "Nr.", "Fisier", "Categorie", "Confidence", "Metoda",
            "Review", "Corectie Manuala", "Reasoning", "Cale Fisier"
        ]

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Freeze first row
    ws.freeze_panes = "A2"

    # === DATA ===
    category_colors = {
        "AGREMENT": "E2EFDA", "AUTORIZATIE_DISTRIBUTIE": "FCE4D6",
        "AVIZ_TEHNIC_SI_AGREMENT": "D6E4F0", "AVIZ_TEHNIC": "DDEBF7",
        "AVIZ_SANITAR": "E2D9F3", "CE": "FFF2CC", "CERTIFICAT_CALITATE": "D5F5E3",
        "CERTIFICAT_GARANTIE": "FADBD8", "CUI": "F5F5F5",
        "DECLARATIE_CONFORMITATE": "DBEEF4", "DECLARATIE_PERFORMANTA": "F0E68C",
        "FISA_TEHNICA": "FDE9D9", "ISO": "E8F5E9", "ALTELE": "F2F2F2",
    }

    flag_colors = {
        "OK": "C6EFCE",
        "REVIEW": "FFEB9C",
        "NEEDS CHECK": "FFC7CE",
    }

    for i, r in enumerate(results):
        row = i + 2
        cat = r["predicted_category"]
        conf = r["confidence"]
        method = r["method"]
        cat_fill = PatternFill(start_color=category_colors.get(cat, "FFFFFF"),
                               end_color=category_colors.get(cat, "FFFFFF"), fill_type="solid")

        # Determine review flag
        if method == "ai" or method == "text_override":
            flag = "NEEDS CHECK"
        elif conf < 0.85:
            flag = "REVIEW"
        elif method == "text_rules" and conf < 0.90:
            flag = "REVIEW"
        else:
            flag = "OK"
        flag_fill = PatternFill(start_color=flag_colors[flag],
                                end_color=flag_colors[flag], fill_type="solid")

        ext = extraction_results[i] if i < len(extraction_results) else {}

        if has_extraction:
            row_data = [
                i + 1,                                          # Nr.
                r["filename"],                                  # Fisier
                cat,                                            # Categorie
                round(conf, 2),                                 # Confidence
                method,                                         # Metoda
                ext.get("nume_document"),                       # Nume Document
                ext.get("numar_pagini"),                        # Numar Pagini
                ext.get("material"),                            # Material
                ext.get("data_expirare"),                       # Data Expirare
                ext.get("companie"),                            # Companie
                ext.get("producator"),                          # Producator
                ext.get("distribuitor"),                        # Distribuitor
                ext.get("adresa_producator"),                   # Adresa Producator
                flag,                                           # Review
                "",                                             # Corectie Manuala
                ext.get("extraction_status", "N/A"),            # Status Extractie
                r["reasoning"][:100],                           # Reasoning
                r["filepath"],                                  # Cale Fisier
            ]
        else:
            row_data = [
                i + 1, r["filename"], cat, round(conf, 2), method,
                flag, "", r["reasoning"][:100], r["filepath"],
            ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Apply category color to Categorie column
        ws.cell(row=row, column=3).fill = cat_fill

        # Apply flag color to Review column
        flag_col = 14 if has_extraction else 6
        ws.cell(row=row, column=flag_col).fill = flag_fill

        # Apply extraction status color
        if has_extraction:
            status = ext.get("extraction_status", "")
            if status == "OK":
                ws.cell(row=row, column=16).fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status.startswith("ERROR"):
                ws.cell(row=row, column=16).fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif status == "NO_TEXT":
                ws.cell(row=row, column=16).fill = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # === COLUMN WIDTHS ===
    if has_extraction:
        widths = {
            'A': 5, 'B': 55, 'C': 32, 'D': 11, 'E': 18,
            'F': 35, 'G': 10, 'H': 45, 'I': 18,
            'J': 30, 'K': 30, 'L': 25, 'M': 40,
            'N': 14, 'O': 22, 'P': 16, 'Q': 50, 'R': 60,
        }
    else:
        widths = {
            'A': 5, 'B': 55, 'C': 32, 'D': 12, 'E': 20,
            'F': 15, 'G': 25, 'H': 50, 'I': 60,
        }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    # Auto-filter
    last_col_letter = chr(ord('A') + len(headers) - 1)
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(results)+1}"

    # === SUMMARY SHEET ===
    ws2 = wb.create_sheet("Sumar")
    ws2.cell(row=1, column=1, value="Categorie").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Nr. Documente").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Procent").font = Font(bold=True)
    i = 1
    for i, (cat, count) in enumerate(sorted(category_counts.items()), 2):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=f"{count/len(results)*100:.1f}%")
    total_row = i + 1
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=len(results)).font = Font(bold=True)

    # Review statistics
    review_row = total_row + 2
    ws2.cell(row=review_row, column=1, value="Statistici Review").font = Font(bold=True, size=12)
    flag_counts = defaultdict(int)
    for r in results:
        conf = r["confidence"]
        method = r["method"]
        if method == "ai" or method == "text_override":
            flag_counts["NEEDS CHECK"] += 1
        elif conf < 0.85 or (method == "text_rules" and conf < 0.90):
            flag_counts["REVIEW"] += 1
        else:
            flag_counts["OK"] += 1

    row = review_row + 1
    ws2.cell(row=row, column=1, value="Metoda").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="Nr.").font = Font(bold=True)
    for m, c in sorted(method_counts.items()):
        row += 1
        ws2.cell(row=row, column=1, value=m)
        ws2.cell(row=row, column=2, value=c)

    row += 2
    ws2.cell(row=row, column=1, value="Review Status").font = Font(bold=True)
    ws2.cell(row=row, column=2, value="Nr.").font = Font(bold=True)
    for flag, c in [("OK", flag_counts.get("OK", 0)),
                    ("REVIEW", flag_counts.get("REVIEW", 0)),
                    ("NEEDS CHECK", flag_counts.get("NEEDS CHECK", 0))]:
        row += 1
        ws2.cell(row=row, column=1, value=flag)
        ws2.cell(row=row, column=2, value=c)
        f_fill = PatternFill(start_color=flag_colors[flag],
                             end_color=flag_colors[flag], fill_type="solid")
        ws2.cell(row=row, column=1).fill = f_fill

    # Extraction statistics
    if has_extraction:
        row += 2
        ws2.cell(row=row, column=1, value="Statistici Extractie").font = Font(bold=True, size=12)
        ext_ok = sum(1 for e in extraction_results if e.get("extraction_status") == "OK")
        ext_err = sum(1 for e in extraction_results if e.get("extraction_status", "").startswith("ERROR"))
        ext_notext = sum(1 for e in extraction_results if e.get("extraction_status") == "NO_TEXT")
        row += 1
        ws2.cell(row=row, column=1, value="OK")
        ws2.cell(row=row, column=2, value=ext_ok)
        ws2.cell(row=row, column=1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        row += 1
        ws2.cell(row=row, column=1, value="Erori AI")
        ws2.cell(row=row, column=2, value=ext_err)
        ws2.cell(row=row, column=1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        row += 1
        ws2.cell(row=row, column=1, value="Fara text (OCR esuat)")
        ws2.cell(row=row, column=2, value=ext_notext)
        ws2.cell(row=row, column=1).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 12

    wb.save(report_path)
    print(f"\nExcel report saved to: {report_path}")


# ============================================================
# TEST EXTRACTION ON SAMPLES
# ============================================================

def test_extraction(input_dir: str, model: str = None, num_samples: int = 3):
    """
    Test extraction on sample documents from each category.
    Useful for evaluating model quality before full run.
    """
    model = model or EXTRACTION_MODEL
    print(f"\n{'='*70}")
    print(f"EXTRACTION TEST — Model: {model}")
    print(f"Samples per category: {num_samples}")
    print(f"{'='*70}")

    # Collect sample files by finding them
    import glob
    all_pdfs = []
    for root, dirs, files in os.walk(input_dir):
        for f in sorted(files):
            if f.lower().endswith('.pdf'):
                all_pdfs.append(os.path.join(root, f))

    if not all_pdfs:
        print(f"No PDFs found in {input_dir}")
        return

    # Classify all, then pick samples from each category
    print(f"\nClassifying {len(all_pdfs)} docs to pick samples...")
    by_category = defaultdict(list)
    for fp in all_pdfs:
        result = classify_document(fp, use_ai=False)
        by_category[result["predicted_category"]].append(fp)

    total_time = 0
    total_ok = 0
    total_tested = 0
    all_results = []

    for cat in sorted(by_category.keys()):
        files = by_category[cat][:num_samples]
        print(f"\n--- {cat} ({len(files)} samples) ---")

        for fp in files:
            filename = os.path.basename(fp)
            total_tested += 1

            start = time.time()
            ext = extract_document_data(fp, cat, model=model)
            elapsed = time.time() - start
            total_time += elapsed

            status = ext.get("extraction_status", "?")
            print(f"\n  [{cat}] {filename[:55]}")
            print(f"  Time: {elapsed:.1f}s | Status: {status}")

            if status == "OK":
                total_ok += 1
                for k in ["nume_document", "material", "data_expirare", "companie",
                           "producator", "distribuitor", "adresa_producator"]:
                    v = ext.get(k)
                    if v:
                        print(f"    {k:20s}: {str(v)[:120]}")
            else:
                print(f"    ERROR: {status}")

            all_results.append({"file": filename, "category": cat, **ext})

    print(f"\n{'='*70}")
    print(f"EXTRACTION TEST RESULTS")
    print(f"{'='*70}")
    print(f"  Model:  {model}")
    print(f"  Tested: {total_tested}")
    print(f"  OK:     {total_ok} ({total_ok/total_tested*100:.0f}%)" if total_tested else "  OK: 0")
    print(f"  Errors: {total_tested - total_ok}")
    print(f"  Time:   {total_time:.1f}s total, {total_time/total_tested:.1f}s avg" if total_tested else "")

    # Save results
    results_path = os.path.join(input_dir, f"test_extraction_{model.replace('/', '_')}.json")
    try:
        with open(results_path, 'w', encoding='utf-8') as f:
            json.dump(all_results, f, ensure_ascii=False, indent=2)
        print(f"\nResults saved to: {results_path}")
    except Exception:
        pass

    return all_results


# ============================================================
# CLI
# ============================================================

if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "help"

    if mode == "eval":
        run_evaluation(use_ai=False)

    elif mode == "classify":
        # Usage: python clasificare_documente_final.py classify /path/to/input [options]
        input_dir = sys.argv[2] if len(sys.argv) > 2 else "."
        output_dir = None
        move = False
        report = None
        do_extract = True
        extract_model = None

        args = sys.argv[3:]
        i = 0
        while i < len(args):
            arg = args[i]
            if arg == "--move":
                move = True
                if not output_dir and i > 0 and not args[0].startswith("--"):
                    output_dir = args[0]
                elif not output_dir:
                    output_dir = input_dir + "_clasificat"
            elif arg == "--report" and i + 1 < len(args):
                report = args[i + 1]
                i += 1
            elif arg == "--no-extract":
                do_extract = False
            elif arg == "--extract-model" and i + 1 < len(args):
                extract_model = args[i + 1]
                i += 1
            elif not arg.startswith("--") and not output_dir:
                output_dir = arg
            i += 1

        if not report:
            report = os.path.join(output_dir or input_dir, "raport_clasificare.xlsx")

        classify_folder(input_dir, output_dir, move, report,
                        do_extract=do_extract, extract_model=extract_model)

    elif mode == "test_extract":
        # Usage: python clasificare_documente_final.py test_extract /path/to/pdfs [--model model_name] [--samples N]
        input_dir = sys.argv[2] if len(sys.argv) > 2 else "."
        model = None
        num_samples = 3

        args = sys.argv[3:]
        i = 0
        while i < len(args):
            if args[i] == "--model" and i + 1 < len(args):
                model = args[i + 1]
                i += 1
            elif args[i] == "--samples" and i + 1 < len(args):
                num_samples = int(args[i + 1])
                i += 1
            i += 1

        test_extraction(input_dir, model=model, num_samples=num_samples)

    elif mode == "compare_models":
        # Usage: python clasificare_documente_final.py compare_models /path/to/pdfs
        input_dir = sys.argv[2] if len(sys.argv) > 2 else "."
        models = [
            "google/gemini-2.0-flash-001",
            "anthropic/claude-3.5-sonnet",
            "openai/gpt-4o-mini",
        ]

        # Allow custom model list
        args = sys.argv[3:]
        if "--models" in args:
            idx = args.index("--models")
            models = args[idx+1].split(",")

        for m in models:
            print(f"\n\n{'#'*70}")
            print(f"# TESTING MODEL: {m}")
            print(f"{'#'*70}")
            test_extraction(input_dir, model=m, num_samples=2)
            time.sleep(1)  # Rate limit between models

    else:
        print(f"""
Document Classification & Extraction System v3
================================================

Usage:
  python {sys.argv[0]} eval                                    # Evaluate classification on training data
  python {sys.argv[0]} classify /path/to/pdfs                  # Classify + extract all PDFs
  python {sys.argv[0]} classify /path/to/pdfs --no-extract     # Classify only (no AI extraction)
  python {sys.argv[0]} classify /input /output --move           # Classify + move to category folders
  python {sys.argv[0]} classify /input --report r.xlsx          # Custom report path
  python {sys.argv[0]} classify /input --extract-model MODEL    # Use specific model for extraction

  python {sys.argv[0]} test_extract /path/to/pdfs              # Test extraction on samples (3 per category)
  python {sys.argv[0]} test_extract /path --model MODEL --samples 5  # Custom model & sample count

  python {sys.argv[0]} compare_models /path/to/pdfs            # Compare 3 models on extraction quality
  python {sys.argv[0]} compare_models /path --models m1,m2,m3  # Compare custom model list

Models available via OpenRouter:
  google/gemini-2.0-flash-001    (fast, cheap - ~$0.001/doc)
  anthropic/claude-3.5-sonnet    (accurate, moderate cost)
  openai/gpt-4o-mini             (good balance)
  google/gemini-2.5-pro-preview-03-25  (best quality, expensive)
""")
