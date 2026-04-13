#!/usr/bin/env python3
"""
Proceseaza Pachete Makyol v2
-----------------------------
Procesează pachete de documente PDF pentru construcții.
Output: Excel cu un sheet per pachet.
Coloane: Nr., Nume document, Numar pagini, Material, Data expirare, Producator, Distribuitor.
Documentele multi-limbă sunt cumulate într-un singur rând.

Utilizare:
    python proceseaza_pachete_v2.py --input "Pachete Makyol" --output Centralizator_Pachete.xlsx
"""

import argparse
import json
import logging
import os
import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# Try OCR imports
try:
    from PIL import Image
    Image.MAX_IMAGE_PIXELS = 300000000
    from pdf2image import convert_from_path
    import pytesseract
    HAS_OCR = True
except ImportError:
    HAS_OCR = False
    logger.warning("OCR libraries not available. Scanned PDFs will have limited text extraction.")

# ── Config ──────────────────────────────────────────────────────────────────

MONTH_MAP_RO = {
    'ianuarie': '01', 'februarie': '02', 'martie': '03', 'aprilie': '04',
    'mai': '05', 'iunie': '06', 'iulie': '07', 'august': '08',
    'septembrie': '09', 'octombrie': '10', 'noiembrie': '11', 'decembrie': '12',
}

MONTH_MAP_EN = {
    'january': '01', 'february': '02', 'march': '03', 'april': '04',
    'may': '05', 'june': '06', 'july': '07', 'august': '08',
    'september': '09', 'october': '10', 'november': '11', 'december': '12',
}

MONTH_MAP_TR = {
    'ocak': '01', 'şubat': '02', 'subat': '02', 'mart': '03', 'nisan': '04',
    'mayıs': '05', 'mayis': '05', 'haziran': '06', 'temmuz': '07', 'ağustos': '08',
    'agustos': '08', 'eylül': '09', 'eylul': '09', 'ekim': '10', 'kasım': '11',
    'kasim': '11', 'aralık': '12', 'aralik': '12',
}

# Combined month map
ALL_MONTHS = {**MONTH_MAP_RO, **MONTH_MAP_EN, **MONTH_MAP_TR}

# Document type patterns (filename-based, then text-based)
DOC_TYPE_PATTERNS = [
    (r'(?i)(?:fi[sș][aă]\s*tehnic[aă]|data\s*sheet|\bFT\b)', 'Fisa tehnica'),
    (r'(?i)ISO\s*9001', 'Certificat ISO 9001'),
    (r'(?i)ISO\s*14001', 'Certificat ISO 14001'),
    (r'(?i)ISO\s*45001', 'Certificat ISO 45001'),
    (r'(?i)ISO\s*50001', 'Certificat ISO 50001'),
    (r'(?i)(?:aviz.*agrement|agrement\s*tehnic|\bAGT\b|\bAT\b\s*\d)', 'Agrement tehnic'),
    (r'(?i)(?:aviz\s*sanitar|\bAVS\b|\bAS\b(?:\s|$))', 'Aviz sanitar'),
    (r'(?i)(?:aviz\s*tehnic|\bAVT\b)', 'Aviz tehnic'),
    (r'(?i)(?:declara[tț]i[ea]\s*de\s*conformitate|\bDC\b(?!\s*-))', 'Declaratie de conformitate'),
    (r'(?i)(?:declara[tț]i[ea]\s*de\s*performan[tț][aă]|\bDOP\b)', 'Declaratie de performanta'),
    (r'(?i)(?:certificat\s*de\s*conformitate|\bCC\b)', 'Certificat de conformitate'),
    (r'(?i)(?:certificat\s*de\s*garan[tț]ie|\bCG\b)', 'Certificat de garantie'),
    (r'(?i)(?:certificat.*3\.1|\bC\s*3\.1\b|\bMTC\b)', 'Certificat 3.1'),
    (r'(?i)(?:certificat.*API)', 'Certificat API'),
    (r'(?i)(?:certificat.*CE\s*PED|CE\s*PED)', 'Certificat CE PED'),
    (r'(?i)(?:certificat.*CPR|CPR)', 'Certificat CPR'),
    (r'(?i)(?:\bCUI\b|certificat.*[îi]nregistrare)', 'Certificat inregistrare (CUI)'),
    (r'(?i)(?:autorizat|autorizare|autorizatie)', 'Autorizatie'),
    (r'(?i)(?:licen[tț][aă])', 'Licenta'),
    (r'(?i)(?:catalog)', 'Catalog'),
    (r'(?i)(?:plan\s*(?:de\s*)?inspec[tț]ie|ITP)', 'Plan inspectie si testare'),
    (r'(?i)(?:procedur[aă])', 'Procedura'),
    (r'(?i)(?:specifica[tț]i)', 'Specificatii'),
    (r'(?i)(?:instruc[tț]iuni)', 'Instructiuni'),
    (r'(?i)(?:garan[tț]ie)', 'Certificat de garantie'),
    (r'(?i)(?:list[aă]\s*proiecte)', 'Lista proiecte'),
]

# Known supplier roles
SUPPLIER_ROLES = {
    'zakprest': {'distributor': 'SC ZAKPREST CONSTRUCT SRL', 'producer': 'TERAPLAST SA'},
    'tehnoworld': {'distributor': 'TEHNO WORLD SRL'},
    'tehno world': {'distributor': 'TEHNO WORLD SRL'},
    'petrouzinex': {'distributor': 'PETROUZINEX SRL'},
    'metal trade': {'distributor': 'METAL TRADE SRL'},
    'metaltrade': {'distributor': 'METAL TRADE SRL'},
    'nimfa': {'distributor': 'NIMFA COM SRL'},
    'titan steel': {'distributor': 'TITAN STEEL TRADING SRL'},
    'innotech': {'distributor': 'INNOTECH SRL'},
    'boma': {'producer': 'BOMA PREFABRICATE SRL'},
}

# Folder name → (producer, distributor) mapping for specific package structures
# Used when folder hierarchy implies the relationship
FOLDER_PRODUCER_MAP = {
    'huayang': ('HEBEI HUAYANG STEEL PIPE CO., LTD.', 'PETROUZINEX SRL'),
    'luzheng': ('WEIFANG LUZHENG INDUSTRY AND TRADE CO., LTD.', 'PETROUZINEX SRL'),
    'steel gate': ('STEEL GATE VALVE CO.', 'PETROUZINEX SRL'),
    'saint gobain': ('SAINT-GOBAIN PAM CANALISATION', 'PETROUZINEX SRL'),
    'saint-gobain': ('SAINT-GOBAIN PAM CANALISATION', 'PETROUZINEX SRL'),
    'bbk': ('BBK VALVE GROUP', None),
    'tianjin kefa': ('TIANJIN KEFA VALVE CO., LTD.', None),
    'akmermer': ('AKMERMER CELIK BORU SAN. VE TIC. A.S.', 'NIMFA COM SRL'),
}

KNOWN_PRODUCERS = {
    'teraplast': 'TERAPLAST SA',
    'valrom': 'VALROM INDUSTRIE SRL',
    'saint gobain': 'SAINT-GOBAIN PAM CANALISATION',
    'saint-gobain': 'SAINT-GOBAIN PAM CANALISATION',
    'huayang': 'HEBEI HUAYANG STEEL PIPE CO., LTD.',
    'luzheng': 'WEIFANG LUZHENG INDUSTRY AND TRADE CO., LTD.',
    'tianjin kefa': 'TIANJIN KEFA VALVE CO., LTD.',
    'tianjin': 'TIANJIN KEFA VALVE CO., LTD.',
    'boma': 'BOMA PREFABRICATE SRL',
    'steel gate': 'STEEL GATE VALVE CO.',
    'bbk': 'BBK VALVE GROUP',
    'nimfa': 'NIMFA COM SRL',
    'akmermer': 'AKMERMER CELIK BORU SAN. VE TIC. A.S.',
    'precon': 'PRECON SRL',
    'bbk': 'BBK VALVE GROUP',
    'innotech': 'INNOTECH VALVES',
    'akmermer': 'AKMERMER',
    'steel gate': 'STEEL GATE',
    'halls': 'HALLS T.',
}

# Multi-language folder pairs
MULTILANG_FOLDERS = {
    'romana': 'ro',
    'akmermer': 'en',
    'en': 'en',
    'eng': 'en',
    'engl': 'en',
    'ro': 'ro',
    'tr': 'tr',
}

# Patterns suggesting a file is a language variant
LANG_VARIANT_PATTERNS = [
    (r'(?i)\b(ro|romana|românesc)\b', 'ro'),
    (r'(?i)\b(en|eng|engl|english)\b', 'en'),
    (r'(?i)\b(tr|turca|turkish)\b', 'tr'),
    (r'(?i)_RO(?:\.|_|$)', 'ro'),
    (r'(?i)_EN(?:\.|_|$)', 'en'),
    (r'(?i)\bro\s*$', 'ro'),
    (r'(?i)\bengl?\s*$', 'en'),
]


# ── PDF Processing ──────────────────────────────────────────────────────────

def get_page_count(pdf_path):
    # Try pdfplumber first
    try:
        with pdfplumber.open(pdf_path) as pdf:
            count = len(pdf.pages)
            if count > 0:
                return count
    except Exception:
        pass
    # Fallback to PyMuPDF (handles scanned PDFs better)
    try:
        import fitz
        doc = fitz.open(str(pdf_path))
        count = len(doc)
        doc.close()
        return count
    except Exception:
        pass
    return 0


def extract_text(pdf_path, max_pages=10):
    text = ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:max_pages]:
                t = page.extract_text() or ""
                if t.strip():
                    text += t + "\n"
    except Exception:
        pass

    if len(text.strip()) < 50 and HAS_OCR:
        try:
            images = convert_from_path(str(pdf_path), dpi=200, first_page=1, last_page=min(max_pages, get_page_count(pdf_path) or 5))
            for img in images:
                t = pytesseract.image_to_string(img, lang='eng')
                if t.strip():
                    text += t + "\n"
        except Exception as e:
            logger.debug(f"OCR failed for {pdf_path.name}: {e}")

    return text


# ── Classification & Extraction ─────────────────────────────────────────────

def classify_document(filename, text=""):
    fname = Path(filename).stem
    for pattern, doc_type in DOC_TYPE_PATTERNS:
        if re.search(pattern, fname):
            return doc_type
    if text:
        for pattern, doc_type in DOC_TYPE_PATTERNS:
            if re.search(pattern, text[:2000]):
                return doc_type
    return "Document tehnic"


def parse_date_to_tuple(raw):
    """Parse a date string to (day, month, year) tuple for comparison. Returns None if unparseable."""
    raw = raw.strip()
    # YYYY-MM-DD
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', raw)
    if m:
        return (int(m.group(3)), int(m.group(2)), int(m.group(1)))
    # DD.MM.YYYY or DD/MM/YYYY
    m = re.match(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', raw)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # DD MonthName YYYY  (any language)
    m = re.match(r'(\d{1,2})\s+(\w+)\s+(\d{4})', raw)
    if m:
        month_str = m.group(2).lower()
        month_num = ALL_MONTHS.get(month_str)
        if month_num:
            return (int(m.group(1)), int(month_num), int(m.group(3)))
    return None


def tuple_to_date_str(t):
    """Convert (day, month, year) tuple to DD.MM.YYYY string."""
    return f"{t[0]:02d}.{t[1]:02d}.{t[2]}"


def extract_expiry_date(text, filename=""):
    """
    Extract expiry date from document text.
    Strategy:
    1. Look for explicit expiry patterns (valabil pana la, valid until, expires)
    2. Collect ALL dates, classify as expiry/emission/other
    3. Prefer the latest future date marked as expiry
    4. Fallback: extract from filename
    """
    if not text and not filename:
        return "-"

    # ── Step 1: Try explicit expiry label patterns (highest confidence) ──
    # These patterns capture dates right after "valid until" type labels
    expiry_label_patterns = [
        # Romanian: "Valabil pana la data: 13 Septembrie 2028"
        r'(?i)valabil[aă]?\s*(?:p[aâ]n[aă]\s*(?:la|[iî]n))\s*(?:data\s*(?:de\s*)?)?:?\s*(\d{1,2}\s+\w+\s+\d{4})',
        r'(?i)valabil[aă]?\s*(?:p[aâ]n[aă]\s*(?:la|[iî]n))\s*(?:data\s*(?:de\s*)?)?:?\s*(\d{1,2}[./]\d{1,2}[./]\d{4})',
        # English: "Valid until: 01.10.2025" / "Expiry date: ..."
        r'(?i)valid\s*(?:until|to|through|till)\s*:?\s*(\d{1,2}[./]\d{1,2}[./]\d{4})',
        r'(?i)valid\s*(?:until|to|through|till)\s*:?\s*(\d{4}-\d{2}-\d{2})',
        r'(?i)valid\s*(?:until|to|through|till)\s*:?\s*(\d{1,2}\s+\w+\s+\d{4})',
        r'(?i)expir[yae]\s*(?:date)?\s*:?\s*(\d{1,2}[./]\d{1,2}[./]\d{4})',
        r'(?i)expir[yae]\s*(?:date)?\s*:?\s*(\d{1,2}\s+\w+\s+\d{4})',
        # Romanian: "pana la" / "pana in"
        r'(?i)p[aâ]n[aă]\s*(?:la|[iî]n)\s*(?:data\s*(?:de\s*)?)?:?\s*(\d{1,2}[./]\d{1,2}[./]\d{4})',
        r'(?i)p[aâ]n[aă]\s*(?:la|[iî]n)\s*(?:data\s*(?:de\s*)?)?:?\s*(\d{1,2}\s+\w+\s+\d{4})',
        # "Data limita" / "Data expirarii"
        r'(?i)data\s*(?:limit[aă]|expir[aă]rii)\s*:?\s*(\d{1,2}[./]\d{1,2}[./]\d{4})',
        r'(?i)data\s*(?:limit[aă]|expir[aă]rii)\s*:?\s*(\d{1,2}\s+\w+\s+\d{4})',
        # "Certificate is valid until"
        r'(?i)certificate\s+is\s+valid\s+(?:until|to)\s*:?\s*(\d{1,2}[./]\d{1,2}[./]\d{4})',
        r'(?i)certificate\s+is\s+valid\s+(?:until|to)\s*:?\s*(\d{1,2}\s+\w+\s+\d{4})',
        r'(?i)certificate\s+is\s+valid\s+(?:until|to)\s*:?\s*(\d{4}-\d{2}-\d{2})',
    ]

    # Search in full text (not just first 500 chars)
    search_text = text or ""
    for p in expiry_label_patterns:
        for m in re.finditer(p, search_text):
            parsed = parse_date_to_tuple(m.group(1))
            if parsed and parsed[2] >= 2024:  # Reasonable expiry year
                return tuple_to_date_str(parsed)

    # ── Step 2: Collect ALL dates and classify ──
    emission_dates = []
    all_dates = []

    # Find dates next to emission labels (to exclude them)
    emission_patterns = [
        r'(?i)(?:data\s*emiter(?:ii|e)|date\s*of\s*issue|issued?\s*(?:on|date)?)\s*:?\s*(\d{1,2}[./]\d{1,2}[./]\d{4})',
        r'(?i)(?:data\s*emiter(?:ii|e)|date\s*of\s*issue|issued?\s*(?:on|date)?)\s*:?\s*(\d{1,2}\s+\w+\s+\d{4})',
        r'(?i)(?:data\s*recertificar[ei]|recertification\s*date)\s*:?\s*(\d{1,2}[./]\d{1,2}[./]\d{4})',
        r'(?i)(?:data\s*recertificar[ei]|recertification\s*date)\s*:?\s*(\d{1,2}\s+\w+\s+\d{4})',
    ]
    for p in emission_patterns:
        for m in re.finditer(p, search_text):
            parsed = parse_date_to_tuple(m.group(1))
            if parsed:
                emission_dates.append(parsed)

    # Collect all numeric dates DD.MM.YYYY or DD/MM/YYYY
    for m in re.finditer(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', search_text):
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= d <= 31 and 1 <= mo <= 12 and 2020 <= y <= 2040:
            all_dates.append((d, mo, y))

    # Collect Romanian/English month name dates
    for m in re.finditer(r'(\d{1,2})\s+(\w+)\s+(\d{4})', search_text):
        month_str = m.group(2).lower()
        month_num = ALL_MONTHS.get(month_str)
        if month_num:
            d, y = int(m.group(1)), int(m.group(3))
            if 1 <= d <= 31 and 2020 <= y <= 2040:
                all_dates.append((d, int(month_num), y))

    # Filter out emission dates, pick the latest remaining
    non_emission = [d for d in all_dates if d not in emission_dates]
    if non_emission:
        # Sort by year, month, day descending — pick the latest (likely expiry)
        non_emission.sort(key=lambda t: (t[2], t[1], t[0]), reverse=True)
        best = non_emission[0]
        if best[2] >= 2024:  # Only return if it's a reasonable expiry
            return tuple_to_date_str(best)

    # If only emission dates exist, still return the latest as fallback
    if all_dates:
        all_dates.sort(key=lambda t: (t[2], t[1], t[0]), reverse=True)
        return tuple_to_date_str(all_dates[0])

    # ── Step 3: Try filename date extraction ──
    if filename:
        # Filename patterns like "2025_2026" or "2028" or "viza 2025"
        m = re.search(r'(\d{1,2})[._](\d{1,2})[._](\d{4})', filename)
        if m:
            parsed = parse_date_to_tuple(f"{m.group(1)}.{m.group(2)}.{m.group(3)}")
            if parsed:
                return tuple_to_date_str(parsed)
        # Just a year in filename
        m = re.search(r'(?:valabil|viza|valid).*?(\d{4})', filename, re.IGNORECASE)
        if m:
            y = int(m.group(1))
            if 2024 <= y <= 2040:
                return f"31.12.{y}"

    return "-"


def extract_material(text, folder_name="", filename=""):
    if not text:
        return infer_material_from_context(folder_name, filename)

    # Priority 1: Explicit labeled material/product fields
    labeled_patterns = [
        r'(?i)(?:denumire\s*produs|product\s*name|produs)\s*:?\s*([^\n]{5,120})',
        r'(?i)(?:material|materiale)\s*:\s*([^\n]{3,80})',
        r'(?i)(?:scope\s*of\s*certification|domeniu(?:l)?\s*de\s*certificare)\s*:?\s*([^\n]{5,150})',
        r'(?i)(?:pentru\s*urm[aă]toarele\s*activit[aă][tț]i)\s*:?\s*\n?\s*([^\n]{5,150})',
    ]
    for p in labeled_patterns:
        m = re.search(p, text[:5000])
        if m:
            result = clean_material_text(m.group(1))
            if result and len(result) >= 5:
                return result

    # Priority 2: Product type patterns (pipes, valves, etc)
    product_patterns = [
        # Pipes
        r'(?i)((?:[TȚtț]evi|[Tt]eav[aă])\s+(?:din\s+)?(?:PEID|PEHD|PE\s*100|PVC|font[aă]\s*ductil[aă]|o[tț]el)[^\n]{0,80})',
        r'(?i)((?:[TȚtț]evi|[Tt]eav[aă])\s+(?:[sș]i|și)\s+fitinguri[^\n]{0,60})',
        # Specific products
        r'(?i)(compensator(?:i)?\s+(?:de\s+)?(?:dilata[tț]ie|metalic)[^\n]{0,60})',
        r'(?i)(c[aă]mine?\s+(?:de\s+)?(?:viz(?:it|a)|beton|prefabricat)[^\n]{0,60})',
        r'(?i)(vane?\s+(?:de\s+)?(?:reglare|sertare|fluture|inchidere)[^\n]{0,60})',
        r'(?i)(garnitur[aă]\s+(?:de\s+)?(?:etan[sș]are|cauciuc|standard)[^\n]{0,60})',
        r'(?i)(tub(?:uri)?\s+(?:din\s+)?(?:o[tț]el|beton|fonta)[^\n]{0,60})',
        r'(?i)(armaturi?\s+[^\n]{5,60})',
    ]
    for p in product_patterns:
        m = re.search(p, text[:5000])
        if m:
            result = clean_material_text(m.group(1))
            if result and len(result) >= 5:
                return result

    # Priority 3: First line of "Fisa Tehnica" style docs
    if re.search(r'(?i)fi[sș][aă]\s*tehnic[aă]', text[:200]):
        # Often the product name is on line 2-4
        lines = text[:500].split('\n')
        for line in lines[1:5]:
            line = line.strip()
            if 10 < len(line) < 120 and not re.search(r'(?i)(data|nr\.|cod|rev|pag)', line):
                return clean_material_text(line)

    return infer_material_from_context(folder_name, filename)


def clean_material_text(raw):
    """Clean and normalize extracted material text."""
    if not raw:
        return None
    result = raw.strip()
    # Remove leading/trailing punctuation
    result = result.strip('.,;:- ')
    # Collapse whitespace
    result = re.sub(r'\s+', ' ', result)
    # Remove garbage patterns
    result = re.sub(r'(?i)^\s*(?:utilizarea|etichetarea|conformitate|nr\.|rev\.|pag\.)\b.*', '', result).strip()
    # Truncate at reasonable point if too long (but at word boundary)
    if len(result) > 100:
        cut = result[:100].rfind(' ')
        if cut > 50:
            result = result[:cut]
    # Don't return if it's just noise
    if len(result) < 3 or re.match(r'^[\d\s.,;:-]+$', result):
        return None
    return result


def infer_material_from_context(folder_name, filename=""):
    folder_lower = folder_name.lower()
    if 'pehd' in folder_lower and 'apa' in folder_lower:
        return "Tevi PEHD PE100 pentru apa"
    if 'pehd' in folder_lower and 'gaz' in folder_lower:
        return "Tevi PEHD PE100 pentru gaz"
    if 'pvc' in folder_lower:
        return "Tevi PVC-KG canalizare"
    if 'fonta' in folder_lower:
        return "Teava fonta ductila"
    if 'gaz' in folder_lower:
        return "Tevi transport gaz"
    if 'vane' in folder_lower or 'valve' in folder_lower:
        return "Vane si armaturi"
    if 'camine' in folder_lower:
        return "Camine prefabricate"
    if 'compensator' in folder_lower:
        return "Compensatori"
    if 'tub' in folder_lower:
        return "Tuburi otel"
    return "-"


def extract_producer(text, folder_path, supplier_name=""):
    """Extract producer from text, folder path, or supplier name."""
    # 1) Try text-based extraction
    if text:
        patterns = [
            r'(?i)produc[aă]tor[:\s]+(?:S\.?C\.?\s*)?([A-Z][A-Z\s.&,]+(?:S\.?R\.?L\.?|S\.?A\.?|CO\.?,?\s*LTD\.?|LTD\.?|GROUP|A\.?S\.?))',
            r'(?i)manufacturer[:\s]+([A-Z][A-Z\s.&,]+(?:S\.?R\.?L\.?|S\.?A\.?|CO\.?,?\s*LTD\.?|LTD\.?))',
            r'(?i)fabricant[:\s]+(?:S\.?C\.?\s*)?([A-Z][A-Z\s.&,]+(?:S\.?R\.?L\.?|S\.?A\.?))',
            # "We, CompanyName, declare..." pattern common in declarations
            r'(?i)we,?\s+([A-Z][A-Za-z\s.&,]+(?:Co\.?,?\s*Ltd\.?|LTD\.?|S\.?R\.?L\.?|S\.?A\.?))',
        ]
        for p in patterns:
            m = re.search(p, text[:5000])
            if m:
                result = m.group(1).strip().rstrip(',. ')
                if 5 < len(result) < 80:
                    return result

    # 2) Infer from FOLDER_PRODUCER_MAP (matches sub-folder names like HUAYANG, LUZHENG)
    combined_path = (folder_path + " " + supplier_name).lower()
    for key, (producer, _distributor) in FOLDER_PRODUCER_MAP.items():
        if key in combined_path:
            return producer

    # 3) Infer from KNOWN_PRODUCERS
    for key, name in KNOWN_PRODUCERS.items():
        if key in combined_path:
            return name

    # 4) Check SUPPLIER_ROLES for producer
    for key, roles in SUPPLIER_ROLES.items():
        if key in combined_path and 'producer' in roles:
            return roles['producer']

    return "-"


def extract_distributor(text, folder_path, supplier_name=""):
    """Extract distributor from text, folder path, or supplier name."""
    # 1) Try text-based extraction
    if text:
        patterns = [
            r'(?i)distribuit(?:or)?[:\s]+(?:S\.?C\.?\s*)?([A-Z][A-Z\s.&,]+(?:S\.?R\.?L\.?|S\.?A\.?))',
            r'(?i)furnizor[:\s]+(?:S\.?C\.?\s*)?([A-Z][A-Z\s.&,]+(?:S\.?R\.?L\.?|S\.?A\.?))',
            r'(?i)distributor[:\s]+(?:S\.?C\.?\s*)?([A-Z][A-Z\s.&,]+(?:S\.?R\.?L\.?|S\.?A\.?))',
            # "autorizam pe SC XYZ SRL ... sa distribuie"
            r'(?i)autoriz[aă]m\s+(?:prin\s+prezenta\s+)?pe\s+(?:S\.?C\.?\s*)?([A-Z][A-Z\s.&,]+(?:S\.?R\.?L\.?|S\.?A\.?))',
        ]
        for p in patterns:
            m = re.search(p, text[:5000])
            if m:
                result = m.group(1).strip().rstrip(',. ')
                if 5 < len(result) < 80:
                    return result

    # 2) Infer from FOLDER_PRODUCER_MAP
    combined_path = (folder_path + " " + supplier_name).lower()
    for key, (_producer, distributor) in FOLDER_PRODUCER_MAP.items():
        if key in combined_path and distributor:
            return distributor

    # 3) Infer from SUPPLIER_ROLES
    for key, roles in SUPPLIER_ROLES.items():
        if key in combined_path and 'distributor' in roles:
            return roles['distributor']

    return "-"


# ── Multi-language Detection ────────────────────────────────────────────────

def detect_language_from_path(filepath):
    parts = Path(filepath).parts
    for part in parts:
        part_lower = part.lower().strip()
        if part_lower in MULTILANG_FOLDERS:
            return MULTILANG_FOLDERS[part_lower]
    stem = Path(filepath).stem
    for pattern, lang in LANG_VARIANT_PATTERNS:
        if re.search(pattern, stem):
            return lang
    return None


def normalize_doc_name(filename):
    stem = Path(filename).stem
    # Remove language indicators and trailing dates/versions
    cleaned = re.sub(r'(?i)[_\s]*(?:ro|romana|en|eng|engl|english|tr|turca)(?:\s|$|_)', ' ', stem)
    cleaned = re.sub(r'(?i)_(?:RO|EN|TR)(?=\.|_|$)', '', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned


def group_multilang_docs(documents):
    """Group documents that are translations of each other."""
    groups = {}
    ungrouped = []

    for doc in documents:
        lang = detect_language_from_path(doc['path'])
        if lang:
            norm_name = normalize_doc_name(doc['filename'])
            key = norm_name.lower()
            if key not in groups:
                groups[key] = {'name': norm_name, 'docs': [], 'langs': []}
            groups[key]['docs'].append(doc)
            groups[key]['langs'].append(lang)
        else:
            ungrouped.append(doc)

    result = []
    for key, group in groups.items():
        if len(group['docs']) > 1:
            # Multiple language versions - merge
            total_pages = sum(d['pages'] for d in group['docs'])
            # Use the Romanian version for metadata, fallback to first
            primary = next((d for d in group['docs'] if detect_language_from_path(d['path']) == 'ro'), group['docs'][0])
            langs_str = "+".join(sorted(set(group['langs'])))
            merged = {
                'filename': primary['filename'],
                'display_name': f"{group['name']} ({langs_str})",
                'pages': total_pages,
                'text': primary.get('text', ''),
                'path': primary['path'],
                'folder_path': primary.get('folder_path', ''),
                'merged': True,
                'lang_count': len(group['docs']),
            }
            result.append(merged)
        else:
            result.append(group['docs'][0])

    result.extend(ungrouped)
    return result


# ── Package Processing ──────────────────────────────────────────────────────

def scan_package(package_dir):
    """Scan a package directory and return structured data."""
    package_dir = Path(package_dir)
    suppliers = []

    for item in sorted(package_dir.iterdir()):
        if not item.is_dir():
            continue
        if item.name.startswith('~$') or item.name.startswith('.'):
            continue

        supplier_name = item.name
        # Check for sub-suppliers (e.g., "3 TR GAZ/HUAYANG")
        sub_dirs = [d for d in item.iterdir() if d.is_dir() and not d.name.startswith('.')]
        has_direct_pdfs = any(f.suffix.lower() == '.pdf' for f in item.iterdir() if f.is_file())

        if sub_dirs and not has_direct_pdfs:
            # This is a category with sub-suppliers
            for sub in sorted(sub_dirs):
                if sub.name.startswith('~$'):
                    continue
                sub_name = f"{supplier_name} / {sub.name}"
                docs = collect_documents(sub, sub_name)
                if docs:
                    suppliers.append({'name': sub_name, 'path': str(sub), 'documents': docs})
        elif sub_dirs and has_direct_pdfs:
            # Has both PDFs at this level and sub-folders
            # Check if sub-folders look like sub-suppliers (contain PDFs)
            sub_has_pdfs = any(
                any(f.suffix.lower() == '.pdf' for f in sub.rglob('*') if f.is_file())
                for sub in sub_dirs
            )
            if sub_has_pdfs and len(sub_dirs) >= 2:
                # Treat as category with sub-suppliers + root PDFs
                # Root PDFs go under the main supplier name
                root_pdfs = [f for f in item.iterdir() if f.is_file() and f.suffix.lower() == '.pdf']
                if root_pdfs:
                    root_docs = []
                    for pdf_path in sorted(root_pdfs):
                        if pdf_path.name.startswith('~$'):
                            continue
                        pages = get_page_count(pdf_path)
                        text = extract_text(pdf_path)
                        root_docs.append({
                            'filename': pdf_path.name,
                            'display_name': pdf_path.stem,
                            'path': str(pdf_path),
                            'rel_path': pdf_path.name,
                            'folder_path': str(item),
                            'supplier': supplier_name,
                            'pages': pages,
                            'text': text,
                            'merged': False,
                        })
                    if root_docs:
                        suppliers.append({'name': supplier_name, 'path': str(item), 'documents': root_docs})
                # Sub-folders as separate suppliers
                for sub in sorted(sub_dirs):
                    if sub.name.startswith('~$'):
                        continue
                    sub_name = f"{supplier_name} / {sub.name}"
                    docs = collect_documents(sub, sub_name)
                    if docs:
                        suppliers.append({'name': sub_name, 'path': str(sub), 'documents': docs})
            else:
                # Simple case: sub-folders are just organization (e.g., APROBARE FURNIZOR)
                docs = collect_documents(item, supplier_name)
                if docs:
                    suppliers.append({'name': supplier_name, 'path': str(item), 'documents': docs})
        else:
            # Simple folder with just PDFs
            docs = collect_documents(item, supplier_name)
            if docs:
                suppliers.append({'name': supplier_name, 'path': str(item), 'documents': docs})

    return suppliers


def collect_documents(folder, supplier_name):
    """Collect all PDF documents from a folder recursively."""
    docs = []
    folder = Path(folder)

    for pdf_path in sorted(folder.rglob('*.pdf')):
        if pdf_path.name.startswith('~$') or 'Thumbs.db' in pdf_path.name:
            continue

        pages = get_page_count(pdf_path)
        text = extract_text(pdf_path)
        rel_path = pdf_path.relative_to(folder)

        docs.append({
            'filename': pdf_path.name,
            'display_name': pdf_path.stem,
            'path': str(pdf_path),
            'rel_path': str(rel_path),
            'folder_path': str(folder),
            'supplier': supplier_name,
            'pages': pages,
            'text': text,
            'merged': False,
        })

    return docs


def process_package(package_dir):
    """Process a single package and return structured results."""
    logger.info(f"Processing package: {Path(package_dir).name}")
    suppliers = scan_package(package_dir)

    results = []
    for supplier in suppliers:
        # Group multi-language documents
        grouped = group_multilang_docs(supplier['documents'])

        processed = []
        for doc in grouped:
            doc_type = classify_document(doc['filename'], doc.get('text', ''))
            expiry = extract_expiry_date(doc.get('text', ''), doc['filename'])
            material = extract_material(doc.get('text', ''), supplier['name'], doc['filename'])
            # Use full doc path for better folder-based inference
            doc_full_path = doc.get('path', doc.get('folder_path', supplier['path']))
            producer = extract_producer(doc.get('text', ''), doc_full_path, supplier['name'])
            distributor = extract_distributor(doc.get('text', ''), doc_full_path, supplier['name'])

            name = doc.get('display_name', doc['filename'])
            if doc.get('merged'):
                name = doc['display_name']

            processed.append({
                'name': name,
                'pages': doc['pages'],
                'material': material,
                'expiry': expiry,
                'producer': producer,
                'distributor': distributor,
                'doc_type': doc_type,
            })

        results.append({
            'supplier': supplier['name'],
            'documents': processed,
        })

    return results


# ── Excel Generation ────────────────────────────────────────────────────────

HEADER_FILL = PatternFill('solid', fgColor='1B4F72')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SUPPLIER_FILL = PatternFill('solid', fgColor='2E86C1')
SUPPLIER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
DATA_FONT = Font(name='Arial', size=9)
DATA_FONT_BOLD = Font(name='Arial', size=9, bold=True)
EVEN_FILL = PatternFill('solid', fgColor='EBF5FB')
BORDER = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7'),
)

COLUMNS = [
    ('Nr.', 5),
    ('Nume document', 45),
    ('Numar pagini', 13),
    ('Material', 40),
    ('Data expirare', 14),
    ('Producator', 30),
    ('Distribuitor', 30),
]


def create_excel(all_packages, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    for pkg_num, pkg_data in all_packages.items():
        sheet_name = f"Pachet {pkg_num}"
        ws = wb.create_sheet(title=sheet_name)

        # Headers
        for col_idx, (header, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        row = 2
        doc_num = 0

        for supplier_data in pkg_data:
            supplier_name = supplier_data['supplier']
            # Supplier separator row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
            cell = ws.cell(row=row, column=1, value=supplier_name)
            cell.font = SUPPLIER_FONT
            cell.fill = SUPPLIER_FILL
            cell.alignment = Alignment(horizontal='left', vertical='center')
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(row=row, column=c).border = BORDER
                ws.cell(row=row, column=c).fill = SUPPLIER_FILL
            row += 1

            for doc in supplier_data['documents']:
                doc_num += 1
                is_even = (doc_num % 2 == 0)
                fill = EVEN_FILL if is_even else PatternFill()

                values = [
                    doc_num,
                    doc['name'],
                    doc['pages'],
                    doc['material'],
                    doc['expiry'],
                    doc['producer'],
                    doc['distributor'],
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    cell.fill = fill
                    if col_idx in (1, 3):  # Nr. and Pages centered
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx == 5:  # Date centered
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(vertical='center', wrap_text=True)

                row += 1

        # Summary row
        total_row = row
        ws.cell(row=total_row, column=1, value="TOTAL").font = DATA_FONT_BOLD
        ws.cell(row=total_row, column=1).border = BORDER
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        # Pages total formula
        pages_col = get_column_letter(3)
        ws.cell(row=total_row, column=3, value=f"=SUM({pages_col}2:{pages_col}{row-1})").font = DATA_FONT_BOLD
        ws.cell(row=total_row, column=3).border = BORDER
        ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='center')
        for c in range(4, len(COLUMNS) + 1):
            ws.cell(row=total_row, column=c).border = BORDER

        # Freeze header
        ws.freeze_panes = 'A2'
        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{row-1}"

    wb.save(output_path)
    logger.info(f"Excel saved: {output_path}")
    return output_path


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Proceseaza Pachete Makyol v2")
    parser.add_argument('--input', '-i', default='Pachete Makyol', help='Director cu pachete')
    parser.add_argument('--output', '-o', default='Centralizator_Pachete.xlsx', help='Fisier Excel output')
    parser.add_argument('--json', '-j', help='Export JSON cu date extrase')
    parser.add_argument('--no-ocr', action='store_true', help='Dezactiveaza OCR')
    args = parser.parse_args()

    if args.no_ocr:
        global HAS_OCR
        HAS_OCR = False

    input_dir = Path(args.input)
    if not input_dir.exists():
        logger.error(f"Input directory not found: {input_dir}")
        sys.exit(1)

    # Discover packages (numbered folders)
    packages = {}
    for item in sorted(input_dir.iterdir()):
        if item.is_dir() and item.name.isdigit():
            packages[int(item.name)] = item

    if not packages:
        logger.error("No numbered package folders found.")
        sys.exit(1)

    logger.info(f"Found {len(packages)} packages: {sorted(packages.keys())}")

    all_results = {}
    for num in sorted(packages.keys()):
        results = process_package(packages[num])
        all_results[num] = results

        total_docs = sum(len(s['documents']) for s in results)
        total_pages = sum(d['pages'] for s in results for d in s['documents'])
        logger.info(f"  Pachet {num}: {len(results)} furnizori, {total_docs} documente, {total_pages} pagini")

    output_path = create_excel(all_results, args.output)

    if args.json:
        # Serialize for JSON (remove text field)
        json_data = {}
        for num, suppliers in all_results.items():
            json_data[str(num)] = []
            for s in suppliers:
                docs = [{k: v for k, v in d.items() if k != 'text'} for d in s['documents']]
                json_data[str(num)].append({'supplier': s['supplier'], 'documents': docs})
        with open(args.json, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        logger.info(f"JSON export: {args.json}")

    # Print summary
    print(f"\n{'='*60}")
    print(f"SUMAR PROCESARE")
    print(f"{'='*60}")
    for num in sorted(all_results.keys()):
        suppliers = all_results[num]
        total_docs = sum(len(s['documents']) for s in suppliers)
        total_pages = sum(d['pages'] for s in suppliers for d in s['documents'])
        print(f"  Pachet {num}: {total_docs:3d} documente, {total_pages:4d} pagini, {len(suppliers)} furnizori")
    total_all = sum(sum(len(s['documents']) for s in suppliers) for suppliers in all_results.values())
    total_pages_all = sum(sum(d['pages'] for s in suppliers for d in s['documents']) for suppliers in all_results.values())
    print(f"{'='*60}")
    print(f"  TOTAL: {total_all} documente, {total_pages_all} pagini")
    print(f"{'='*60}")
    print(f"\nOutput: {output_path}")


if __name__ == '__main__':
    main()
