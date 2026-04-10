"""End-to-end integration tests for compliance report generation system."""
import pytest
import time
from io import BytesIO
from datetime import date, timedelta
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.main import app
from src.database import get_db, init_db


@pytest.fixture(scope="module")
def client():
    """Create test client for API testing."""
    # Initialize database
    init_db()

    # Create test client
    client = TestClient(app)

    # Generate sample data before tests
    response = client.post("/api/sample-data/generate")
    assert response.status_code == 200

    yield client


class TestEndToEndReportGeneration:
    """End-to-end tests for all report types and formats."""

    def test_supplier_summary_pdf(self, client):
        """Test Supplier Summary report in PDF format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "supplier_summary",
                "format": "pdf",
                "preparer": "Integration Test User"
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert "Content-Disposition" in response.headers
        assert "supplier_summary" in response.headers["Content-Disposition"]

        # Verify file size
        content = response.content
        assert len(content) > 0
        assert content.startswith(b'%PDF')

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_supplier_summary_excel(self, client):
        """Test Supplier Summary report in Excel format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "supplier_summary",
                "format": "excel",
                "preparer": "Integration Test User"
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "supplier_summary" in response.headers["Content-Disposition"]

        # Verify file size and validity
        content = response.content
        assert len(content) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(content))
        assert wb is not None
        assert len(wb.worksheets) > 0

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_document_inventory_pdf(self, client):
        """Test Document Inventory report in PDF format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "document_inventory",
                "format": "pdf",
                "preparer": "Integration Test User"
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert "document_inventory" in response.headers["Content-Disposition"]

        # Verify file size
        content = response.content
        assert len(content) > 0
        assert content.startswith(b'%PDF')

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_document_inventory_excel(self, client):
        """Test Document Inventory report in Excel format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "document_inventory",
                "format": "excel",
                "preparer": "Integration Test User"
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "document_inventory" in response.headers["Content-Disposition"]

        # Verify file size and validity
        content = response.content
        assert len(content) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(content))
        assert wb is not None

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_expiring_certificates_pdf(self, client):
        """Test Expiring Certificates report in PDF format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "expiring_certificates",
                "format": "pdf",
                "preparer": "Integration Test User",
                "days_threshold": 30
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert "expiring_certificates" in response.headers["Content-Disposition"]

        # Verify file size
        content = response.content
        assert len(content) > 0
        assert content.startswith(b'%PDF')

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_expiring_certificates_excel(self, client):
        """Test Expiring Certificates report in Excel format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "expiring_certificates",
                "format": "excel",
                "preparer": "Integration Test User",
                "days_threshold": 30
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "expiring_certificates" in response.headers["Content-Disposition"]

        # Verify file size and validity
        content = response.content
        assert len(content) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(content))
        assert wb is not None

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_missing_documents_pdf(self, client):
        """Test Missing Documents report in PDF format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "missing_documents",
                "format": "pdf",
                "preparer": "Integration Test User"
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert "missing_documents" in response.headers["Content-Disposition"]

        # Verify file size
        content = response.content
        assert len(content) > 0
        assert content.startswith(b'%PDF')

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_missing_documents_excel(self, client):
        """Test Missing Documents report in Excel format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "missing_documents",
                "format": "excel",
                "preparer": "Integration Test User"
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "missing_documents" in response.headers["Content-Disposition"]

        # Verify file size and validity
        content = response.content
        assert len(content) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(content))
        assert wb is not None

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_full_audit_pdf(self, client):
        """Test Full Audit report in PDF format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "full_audit",
                "format": "pdf",
                "preparer": "Integration Test User"
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert "full_audit" in response.headers["Content-Disposition"]

        # Verify file size
        content = response.content
        assert len(content) > 0
        assert content.startswith(b'%PDF')

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"

    def test_full_audit_excel(self, client):
        """Test Full Audit report in Excel format."""
        start_time = time.time()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "full_audit",
                "format": "excel",
                "preparer": "Integration Test User"
            }
        )

        generation_time = time.time() - start_time

        # Verify response
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "full_audit" in response.headers["Content-Disposition"]

        # Verify file size and validity
        content = response.content
        assert len(content) > 0

        # Verify it's a valid Excel file with multiple worksheets
        wb = load_workbook(BytesIO(content))
        assert wb is not None
        assert len(wb.worksheets) >= 4  # Multiple worksheets for full audit

        # Verify generation time
        assert generation_time < 5, f"Generation took {generation_time:.2f}s (should be < 5s)"


class TestFilteringFunctionality:
    """Test filtering capabilities for reports."""

    def test_filter_by_supplier(self, client):
        """Test filtering reports by supplier ID."""
        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "document_inventory",
                "format": "pdf",
                "preparer": "Integration Test User",
                "supplier_id": 1
            }
        )

        # Verify response
        assert response.status_code == 200
        assert len(response.content) > 0
        assert response.content.startswith(b'%PDF')

    def test_filter_by_date_range(self, client):
        """Test filtering reports by date range."""
        today = date.today()
        date_from = (today - timedelta(days=30)).isoformat()
        date_to = (today + timedelta(days=60)).isoformat()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "document_inventory",
                "format": "excel",
                "preparer": "Integration Test User",
                "date_from": date_from,
                "date_to": date_to
            }
        )

        # Verify response
        assert response.status_code == 200
        assert len(response.content) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(response.content))
        assert wb is not None

    def test_filter_by_status(self, client):
        """Test filtering reports by document status."""
        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "document_inventory",
                "format": "pdf",
                "preparer": "Integration Test User",
                "status": "valid"
            }
        )

        # Verify response
        assert response.status_code == 200
        assert len(response.content) > 0
        assert response.content.startswith(b'%PDF')

    def test_filter_with_multiple_parameters(self, client):
        """Test filtering with multiple parameters combined."""
        today = date.today()

        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "document_inventory",
                "format": "excel",
                "preparer": "Integration Test User",
                "supplier_id": 1,
                "status": "valid",
                "date_from": today.isoformat(),
                "date_to": (today + timedelta(days=365)).isoformat()
            }
        )

        # Verify response
        assert response.status_code == 200
        assert len(response.content) > 0

        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(response.content))
        assert wb is not None

    def test_expiring_certificates_days_threshold(self, client):
        """Test expiring certificates report with custom days threshold."""
        # Test with 60 days threshold
        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "expiring_certificates",
                "format": "pdf",
                "preparer": "Integration Test User",
                "days_threshold": 60
            }
        )

        # Verify response
        assert response.status_code == 200
        assert len(response.content) > 0
        assert response.content.startswith(b'%PDF')


class TestAPIHealthAndSampleData:
    """Test API health endpoints and sample data generation."""

    def test_root_endpoint(self, client):
        """Test root endpoint returns API information."""
        response = client.get("/")

        assert response.status_code == 200
        data = response.json()
        assert "app" in data
        assert "version" in data
        assert "status" in data
        assert data["status"] == "running"

    def test_health_check(self, client):
        """Test health check endpoint."""
        response = client.get("/health")

        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "healthy"

    def test_sample_data_generation(self, client):
        """Test sample data generation endpoint."""
        response = client.post("/api/sample-data/generate")

        assert response.status_code == 200
        data = response.json()
        assert "message" in data
        assert "suppliers_created" in data
        assert "documents_created" in data
        assert data["suppliers_created"] > 0
        assert data["documents_created"] > 0


class TestErrorHandling:
    """Test error handling and validation."""

    def test_invalid_report_type(self, client):
        """Test error handling for invalid report type."""
        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "invalid_report",
                "format": "pdf",
                "preparer": "Test User"
            }
        )

        # Should return validation error
        assert response.status_code == 422  # Pydantic validation error

    def test_invalid_format(self, client):
        """Test error handling for invalid format."""
        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "supplier_summary",
                "format": "invalid_format",
                "preparer": "Test User"
            }
        )

        # Should return validation error
        assert response.status_code == 422  # Pydantic validation error

    def test_missing_preparer(self, client):
        """Test error handling for missing preparer field."""
        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "supplier_summary",
                "format": "pdf"
            }
        )

        # Should return validation error
        assert response.status_code == 422  # Pydantic validation error

    def test_invalid_days_threshold(self, client):
        """Test error handling for invalid days threshold."""
        response = client.post(
            "/api/reports/generate",
            json={
                "report_type": "expiring_certificates",
                "format": "pdf",
                "preparer": "Test User",
                "days_threshold": 500  # Exceeds maximum of 365
            }
        )

        # Should return validation error
        assert response.status_code == 422  # Pydantic validation error


class TestE2EAllReportsSequentially:
    """End-to-end test that generates all report types sequentially."""

    def test_e2e_all_reports(self, client):
        """
        Comprehensive end-to-end test generating all reports.
        This verifies the complete workflow from data generation to report download.
        """
        # Define all report configurations
        report_configs = [
            ("supplier_summary", "pdf"),
            ("supplier_summary", "excel"),
            ("document_inventory", "pdf"),
            ("document_inventory", "excel"),
            ("expiring_certificates", "pdf"),
            ("expiring_certificates", "excel"),
            ("missing_documents", "pdf"),
            ("missing_documents", "excel"),
            ("full_audit", "pdf"),
            ("full_audit", "excel"),
        ]

        total_start_time = time.time()
        results = []

        for report_type, format_type in report_configs:
            start_time = time.time()

            # Generate report
            response = client.post(
                "/api/reports/generate",
                json={
                    "report_type": report_type,
                    "format": format_type,
                    "preparer": "E2E Test User"
                }
            )

            generation_time = time.time() - start_time

            # Verify response
            assert response.status_code == 200, f"Failed for {report_type} in {format_type} format"
            assert len(response.content) > 0, f"Empty content for {report_type} in {format_type} format"

            # Verify generation time
            assert generation_time < 5, f"{report_type} ({format_type}) took {generation_time:.2f}s (should be < 5s)"

            results.append({
                "report_type": report_type,
                "format": format_type,
                "size_bytes": len(response.content),
                "generation_time": generation_time,
                "status": "success"
            })

        total_time = time.time() - total_start_time

        # Log results summary
        print("\n" + "="*80)
        print("E2E Test Results Summary")
        print("="*80)
        for result in results:
            print(f"{result['report_type']:25} | {result['format']:6} | "
                  f"{result['size_bytes']:8} bytes | {result['generation_time']:.3f}s")
        print("="*80)
        print(f"Total time for all reports: {total_time:.2f}s")
        print(f"Average time per report: {total_time / len(results):.2f}s")
        print("="*80)

        # Verify all reports were successful
        assert len(results) == 10, "Should have generated 10 reports (5 types × 2 formats)"
        assert all(r["status"] == "success" for r in results), "All reports should succeed"
