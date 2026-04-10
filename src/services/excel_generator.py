"""Excel report generator service using openpyxl."""
from io import BytesIO
from typing import Optional, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelGenerator:
    """Service for generating Excel reports with Makyol branding and formatting."""

    # Brand colors (matching PDF template)
    MAKYOL_PRIMARY = "1B4F72"  # Dark blue
    MAKYOL_SECONDARY = "2E86C1"  # Light blue
    MAKYOL_ACCENT = "5DADE2"  # Accent blue
    HEADER_BG = "F8F9F9"  # Light grey for table headers
    WHITE = "FFFFFF"

    def __init__(self):
        """Initialize Excel generator with Makyol branding."""
        self.workbook = None
        self.current_sheet = None

    def create_workbook(self, sheet_name: str = "Report") -> Workbook:
        """
        Create a new Excel workbook with initial worksheet.

        Args:
            sheet_name: Name for the first worksheet

        Returns:
            Workbook instance
        """
        self.workbook = Workbook()
        self.current_sheet = self.workbook.active
        self.current_sheet.title = sheet_name
        return self.workbook

    def add_worksheet(self, sheet_name: str) -> Worksheet:
        """
        Add a new worksheet to the workbook.

        Args:
            sheet_name: Name for the new worksheet

        Returns:
            Worksheet instance
        """
        if self.workbook is None:
            self.create_workbook(sheet_name)
        else:
            self.current_sheet = self.workbook.create_sheet(title=sheet_name)
        return self.current_sheet

    def apply_header_style(
        self,
        worksheet: Worksheet,
        row_number: int = 1,
        num_columns: int = 1
    ) -> None:
        """
        Apply Makyol header styling to a row.

        Args:
            worksheet: Worksheet to apply styling to
            row_number: Row number to style (default: 1)
            num_columns: Number of columns in the header row
        """
        # Define header styles
        header_font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color=self.MAKYOL_PRIMARY
        )
        header_fill = PatternFill(
            start_color=self.HEADER_BG,
            end_color=self.HEADER_BG,
            fill_type='solid'
        )
        header_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        header_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='medium', color=self.MAKYOL_PRIMARY)
        )

        # Apply styles to header row
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=row_number, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

    def apply_data_style(
        self,
        worksheet: Worksheet,
        start_row: int = 2,
        num_rows: int = 1,
        num_columns: int = 1
    ) -> None:
        """
        Apply standard data styling to table rows.

        Args:
            worksheet: Worksheet to apply styling to
            start_row: First data row (default: 2)
            num_rows: Number of data rows
            num_columns: Number of columns
        """
        # Define data styles
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=False
        )
        data_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # Alternating row colors
        row_color_1 = PatternFill(
            start_color=self.WHITE,
            end_color=self.WHITE,
            fill_type='solid'
        )
        row_color_2 = PatternFill(
            start_color=self.HEADER_BG,
            end_color=self.HEADER_BG,
            fill_type='solid'
        )

        # Apply styles to data rows
        for row in range(start_row, start_row + num_rows):
            # Alternate row colors
            fill = row_color_1 if (row - start_row) % 2 == 0 else row_color_2

            for col in range(1, num_columns + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.fill = fill
                cell.alignment = data_alignment
                cell.border = data_border

    def auto_adjust_column_width(
        self,
        worksheet: Worksheet,
        min_width: int = 10,
        max_width: int = 50
    ) -> None:
        """
        Auto-adjust column widths based on content.

        Args:
            worksheet: Worksheet to adjust
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def apply_auto_filter(
        self,
        worksheet: Worksheet,
        start_row: int = 1,
        end_row: Optional[int] = None,
        end_column: Optional[int] = None
    ) -> None:
        """
        Apply auto-filter to a table.

        Args:
            worksheet: Worksheet to apply filter to
            start_row: Header row number (default: 1)
            end_row: Last row of data (if None, uses last row in sheet)
            end_column: Last column of data (if None, uses last column in sheet)
        """
        if end_row is None:
            end_row = worksheet.max_row
        if end_column is None:
            end_column = worksheet.max_column

        # Apply auto-filter
        filter_range = f"A{start_row}:{get_column_letter(end_column)}{end_row}"
        worksheet.auto_filter.ref = filter_range

    def freeze_panes(
        self,
        worksheet: Worksheet,
        row: int = 2,
        column: int = 1
    ) -> None:
        """
        Freeze header row and optionally first columns.

        Args:
            worksheet: Worksheet to freeze panes on
            row: Row number to freeze above (default: 2 - freezes header row)
            column: Column number to freeze left of (default: 1 - no column freeze)
        """
        freeze_cell = f"{get_column_letter(column)}{row}"
        worksheet.freeze_panes = freeze_cell

    def add_title_section(
        self,
        worksheet: Worksheet,
        title: str,
        subtitle: Optional[str] = None,
        report_info: Optional[dict] = None
    ) -> int:
        """
        Add a title section to the worksheet.

        Args:
            worksheet: Worksheet to add title to
            title: Main title text
            subtitle: Optional subtitle text
            report_info: Optional dictionary of report metadata

        Returns:
            The next available row number after the title section
        """
        current_row = 1

        # Company name (Makyol branding)
        worksheet.cell(row=current_row, column=1, value="MAKYOL")
        worksheet.cell(row=current_row, column=1).font = Font(
            name='Calibri',
            size=18,
            bold=True,
            color=self.MAKYOL_PRIMARY
        )
        current_row += 1

        # Subtitle
        if subtitle:
            worksheet.cell(row=current_row, column=1, value=subtitle)
            worksheet.cell(row=current_row, column=1).font = Font(
                name='Calibri',
                size=12,
                color=self.MAKYOL_SECONDARY
            )
            current_row += 1

        # Main title
        worksheet.cell(row=current_row, column=1, value=title)
        worksheet.cell(row=current_row, column=1).font = Font(
            name='Calibri',
            size=14,
            bold=True,
            color=self.MAKYOL_PRIMARY
        )
        current_row += 2

        # Report info
        if report_info:
            for key, value in report_info.items():
                worksheet.cell(row=current_row, column=1, value=f"{key}:")
                worksheet.cell(row=current_row, column=1).font = Font(
                    name='Calibri',
                    size=10,
                    bold=True,
                    color=self.MAKYOL_SECONDARY
                )
                worksheet.cell(row=current_row, column=2, value=str(value))
                worksheet.cell(row=current_row, column=2).font = Font(
                    name='Calibri',
                    size=10
                )
                current_row += 1
            current_row += 1

        return current_row

    def add_footer(
        self,
        worksheet: Worksheet,
        preparer: Optional[str] = None
    ) -> None:
        """
        Add footer information to the worksheet.

        Args:
            worksheet: Worksheet to add footer to
            preparer: Name of the person who generated the report
        """
        generation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        footer_text = f"Generated on {generation_date}"
        if preparer:
            footer_text += f" | Prepared by: {preparer}"
        footer_text += " | Makyol Compliance Management System"

        # Add footer to worksheet header/footer
        worksheet.oddFooter.center.text = footer_text
        worksheet.oddFooter.center.size = 8

    def create_table(
        self,
        worksheet: Worksheet,
        data: List[List],
        start_row: int = 1,
        start_column: int = 1,
        apply_formatting: bool = True,
        apply_filters: bool = True,
        freeze_header: bool = True
    ) -> int:
        """
        Create a formatted table in the worksheet.

        Args:
            worksheet: Worksheet to add table to
            data: Table data (including header row)
            start_row: Starting row number (default: 1)
            start_column: Starting column number (default: 1)
            apply_formatting: Whether to apply header and data styling
            apply_filters: Whether to apply auto-filter
            freeze_header: Whether to freeze header row

        Returns:
            The next available row number after the table
        """
        if not data:
            return start_row

        num_rows = len(data)
        num_columns = len(data[0]) if data else 0

        # Write data to worksheet
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                cell = worksheet.cell(
                    row=start_row + row_idx,
                    column=start_column + col_idx,
                    value=value
                )

        # Apply formatting
        if apply_formatting:
            # Apply header style to first row
            self.apply_header_style(
                worksheet,
                row_number=start_row,
                num_columns=num_columns
            )

            # Apply data style to remaining rows
            if num_rows > 1:
                self.apply_data_style(
                    worksheet,
                    start_row=start_row + 1,
                    num_rows=num_rows - 1,
                    num_columns=num_columns
                )

            # Auto-adjust column widths
            self.auto_adjust_column_width(worksheet)

        # Apply auto-filter
        if apply_filters:
            self.apply_auto_filter(
                worksheet,
                start_row=start_row,
                end_row=start_row + num_rows - 1,
                end_column=start_column + num_columns - 1
            )

        # Freeze header row
        if freeze_header:
            self.freeze_panes(worksheet, row=start_row + 1)

        return start_row + num_rows

    def save_to_buffer(self) -> BytesIO:
        """
        Save the workbook to a BytesIO buffer.

        Returns:
            BytesIO buffer containing the Excel file
        """
        if self.workbook is None:
            raise ValueError("No workbook created. Call create_workbook() first.")

        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_basic_report(
        self,
        report_title: str,
        data: List[List],
        sheet_name: str = "Report",
        preparer: Optional[str] = None,
        report_info: Optional[dict] = None
    ) -> BytesIO:
        """
        Generate a basic Excel report with single table.

        Args:
            report_title: Title of the report
            data: Table data (including header row)
            sheet_name: Name for the worksheet
            preparer: Name of the person generating the report
            report_info: Dictionary with report metadata

        Returns:
            BytesIO buffer containing the Excel file

        Example:
            data = [
                ['Header1', 'Header2', 'Header3'],
                ['Row1Col1', 'Row1Col2', 'Row1Col3'],
                ['Row2Col1', 'Row2Col2', 'Row2Col3']
            ]
        """
        # Create workbook
        self.create_workbook(sheet_name=sheet_name)

        # Add title section if report_info provided
        start_row = 1
        if report_info or report_title:
            start_row = self.add_title_section(
                self.current_sheet,
                title=report_title,
                subtitle="Compliance Management System",
                report_info=report_info
            )

        # Create table
        self.create_table(
            self.current_sheet,
            data=data,
            start_row=start_row,
            apply_formatting=True,
            apply_filters=True,
            freeze_header=True
        )

        # Add footer
        self.add_footer(self.current_sheet, preparer)

        # Save to buffer
        return self.save_to_buffer()
