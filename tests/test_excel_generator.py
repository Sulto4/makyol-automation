"""Tests for Excel generator service."""
import pytest
from io import BytesIO
from datetime import datetime, timedelta
from openpyxl import load_workbook

from src.services.excel_generator import ExcelGenerator


def test_supplier_summary_excel():
    """Test generation of Supplier Summary Excel report."""
    # Initialize generator
    generator = ExcelGenerator()

    # Sample supplier summary data
    # Table format: Supplier Name | Status | Document Count | Expiring Soon
    supplier_data = [
        {
            'name': 'ABC Construction Ltd.',
            'status': 'active',
            'document_count': 12,
            'expiring_soon': 2
        },
        {
            'name': 'XYZ Engineering Co.',
            'status': 'active',
            'document_count': 8,
            'expiring_soon': 0
        },
        {
            'name': 'BuildTech Industries',
            'status': 'pending',
            'document_count': 5,
            'expiring_soon': 1
        }
    ]

    # Generate report
    excel_buffer = generator.generate_supplier_summary_excel(
        supplier_data=supplier_data,
        preparer="Test User"
    )

    # Verify buffer is created and has content
    assert isinstance(excel_buffer, BytesIO)
    assert len(excel_buffer.getvalue()) > 0

    # Verify it's a valid Excel file
    wb = load_workbook(excel_buffer)
    assert wb is not None
    assert len(wb.worksheets) > 0

    # Verify auto-filter is applied (sortable columns)
    ws = wb.active
    assert ws.auto_filter is not None
    assert ws.auto_filter.ref is not None


def test_document_inventory_excel():
    """Test generation of Document Inventory Excel report."""
    # Initialize generator
    generator = ExcelGenerator()

    # Sample document inventory data
    # Table format: Supplier | Document Type | Status | Validity Date | Certificate #
    document_data = [
        {
            'supplier_name': 'ABC Construction Ltd.',
            'document_type': 'ISO 9001 Certification',
            'status': 'valid',
            'validity_date': '2025-06-30',
            'certificate_number': 'ISO-9001-2024-001'
        },
        {
            'supplier_name': 'ABC Construction Ltd.',
            'document_type': 'Tax Clearance',
            'status': 'expiring',
            'validity_date': '2026-05-15',
            'certificate_number': 'TAX-2024-456'
        },
        {
            'supplier_name': 'XYZ Engineering Co.',
            'document_type': 'ISO 9001 Certification',
            'status': 'valid',
            'validity_date': '2027-01-20',
            'certificate_number': 'ISO-9001-2024-002'
        },
        {
            'supplier_name': 'BuildTech Industries',
            'document_type': 'Insurance Certificate',
            'status': 'expired',
            'validity_date': '2025-12-31',
            'certificate_number': 'INS-2024-789'
        }
    ]

    # Generate report
    excel_buffer = generator.generate_document_inventory_excel(
        document_data=document_data,
        preparer="Test User"
    )

    # Verify buffer is created and has content
    assert isinstance(excel_buffer, BytesIO)
    assert len(excel_buffer.getvalue()) > 0

    # Verify it's a valid Excel file
    wb = load_workbook(excel_buffer)
    assert wb is not None
    assert len(wb.worksheets) > 0

    # Verify auto-filter is applied (sortable columns)
    ws = wb.active
    assert ws.auto_filter is not None
    assert ws.auto_filter.ref is not None


def test_expiring_certificates_excel():
    """Test generation of Expiring Certificates Excel report."""
    # Initialize generator
    generator = ExcelGenerator()

    # Sample expiring certificates data
    # Table format: Supplier | Document Type | Validity Date | Days Until Expiry | Certificate #
    certificate_data = [
        {
            'supplier_name': 'ABC Construction Ltd.',
            'document_type': 'Tax Clearance',
            'validity_date': '2026-05-15',
            'days_until_expiry': 35,
            'certificate_number': 'TAX-2024-456'
        },
        {
            'supplier_name': 'BuildTech Industries',
            'document_type': 'Insurance Certificate',
            'validity_date': '2026-04-25',
            'days_until_expiry': 15,
            'certificate_number': 'INS-2024-789'
        },
        {
            'supplier_name': 'XYZ Engineering Co.',
            'document_type': 'ISO 9001 Certification',
            'validity_date': '2026-05-01',
            'days_until_expiry': 21,
            'certificate_number': 'ISO-9001-2024-002'
        }
    ]

    # Generate report
    excel_buffer = generator.generate_expiring_certificates_excel(
        certificate_data=certificate_data,
        preparer="Test User"
    )

    # Verify buffer is created and has content
    assert isinstance(excel_buffer, BytesIO)
    assert len(excel_buffer.getvalue()) > 0

    # Verify it's a valid Excel file
    wb = load_workbook(excel_buffer)
    assert wb is not None
    assert len(wb.worksheets) > 0

    # Verify auto-filter is applied (sortable columns)
    ws = wb.active
    assert ws.auto_filter is not None
    assert ws.auto_filter.ref is not None


def test_missing_documents_excel():
    """Test generation of Missing Documents Excel report."""
    # Initialize generator
    generator = ExcelGenerator()

    # Sample missing documents data
    # Table format: Supplier | Document Type | Required | Status
    missing_data = [
        {
            'supplier_name': 'ABC Construction Ltd.',
            'document_type': 'Environmental Compliance Certificate',
            'required': 'Always Required',
            'status': 'Missing'
        },
        {
            'supplier_name': 'BuildTech Industries',
            'document_type': 'Safety Certification',
            'required': 'Always Required',
            'status': 'Missing'
        },
        {
            'supplier_name': 'XYZ Engineering Co.',
            'document_type': 'Quality Assurance Certificate',
            'required': 'Always Required',
            'status': 'Missing'
        },
        {
            'supplier_name': 'BuildTech Industries',
            'document_type': 'Tax Clearance',
            'required': 'Always Required',
            'status': 'Missing'
        }
    ]

    # Generate report
    excel_buffer = generator.generate_missing_documents_excel(
        missing_data=missing_data,
        preparer="Test User"
    )

    # Verify buffer is created and has content
    assert isinstance(excel_buffer, BytesIO)
    assert len(excel_buffer.getvalue()) > 0

    # Verify it's a valid Excel file
    wb = load_workbook(excel_buffer)
    assert wb is not None
    assert len(wb.worksheets) > 0

    # Verify auto-filter is applied (sortable columns)
    ws = wb.active
    assert ws.auto_filter is not None
    assert ws.auto_filter.ref is not None


def test_full_audit_excel():
    """Test generation of Full Audit Excel report."""
    # Initialize generator
    generator = ExcelGenerator()

    # Sample data for full audit (combines all report types)
    # Supplier summary data
    supplier_data = [
        {
            'name': 'ABC Construction Ltd.',
            'status': 'active',
            'document_count': 12,
            'expiring_soon': 2
        },
        {
            'name': 'XYZ Engineering Co.',
            'status': 'active',
            'document_count': 8,
            'expiring_soon': 0
        },
        {
            'name': 'BuildTech Industries',
            'status': 'pending',
            'document_count': 5,
            'expiring_soon': 1
        }
    ]

    # Document inventory data
    document_data = [
        {
            'supplier_name': 'ABC Construction Ltd.',
            'document_type': 'ISO 9001 Certification',
            'status': 'valid',
            'validity_date': '2025-06-30',
            'certificate_number': 'ISO-9001-2024-001'
        },
        {
            'supplier_name': 'ABC Construction Ltd.',
            'document_type': 'Tax Clearance',
            'status': 'expiring',
            'validity_date': '2026-05-15',
            'certificate_number': 'TAX-2024-456'
        },
        {
            'supplier_name': 'XYZ Engineering Co.',
            'document_type': 'ISO 9001 Certification',
            'status': 'valid',
            'validity_date': '2027-01-20',
            'certificate_number': 'ISO-9001-2024-002'
        }
    ]

    # Expiring certificates data
    certificate_data = [
        {
            'supplier_name': 'ABC Construction Ltd.',
            'document_type': 'Tax Clearance',
            'validity_date': '2026-05-15',
            'days_until_expiry': 35,
            'certificate_number': 'TAX-2024-456'
        },
        {
            'supplier_name': 'BuildTech Industries',
            'document_type': 'Insurance Certificate',
            'validity_date': '2026-04-25',
            'days_until_expiry': 15,
            'certificate_number': 'INS-2024-789'
        }
    ]

    # Missing documents data
    missing_data = [
        {
            'supplier_name': 'ABC Construction Ltd.',
            'document_type': 'Environmental Compliance Certificate',
            'required': 'Always Required',
            'status': 'Missing'
        },
        {
            'supplier_name': 'BuildTech Industries',
            'document_type': 'Safety Certification',
            'required': 'Always Required',
            'status': 'Missing'
        }
    ]

    # Generate full audit report
    excel_buffer = generator.generate_full_audit_excel(
        supplier_data=supplier_data,
        document_data=document_data,
        certificate_data=certificate_data,
        missing_data=missing_data,
        preparer="Test User"
    )

    # Verify buffer is created and has content
    assert isinstance(excel_buffer, BytesIO)
    assert len(excel_buffer.getvalue()) > 0

    # Verify it's a valid Excel file
    wb = load_workbook(excel_buffer)
    assert wb is not None

    # Verify multiple worksheets exist (one for each section)
    assert len(wb.worksheets) >= 4  # At least 4 sections (Executive Summary + 4 data sheets)

    # Verify auto-filter is applied to data worksheets (not Executive Summary)
    data_sheets = ['Supplier Summary', 'Document Inventory', 'Expiring Certificates', 'Missing Documents']
    for sheet_name in data_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            assert ws.auto_filter is not None
            assert ws.auto_filter.ref is not None
